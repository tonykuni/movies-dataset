"""VPL-XLS · transcript-review Excel template (reduce manual classification).

Produces the 'Transcript_Review' table (the Power-Automate flow's Excel) with
classification columns + an Action checkbox column and category data-validation,
so participants can be checked / categories picked with minimal typing.
"""
# ===== [VIA:ACCEL-BRIDGE:v0100] SuperAccel 加速器橋(全引擎導入令 2026-08-18;graceful 零行為變更) =====
try:
    import sys as _sa_sys
    from pathlib import Path as _sa_Path
    _sa_p = _sa_Path(__file__).resolve()
    while _sa_p.parent != _sa_p:
        if (_sa_p / "supportive modules" / "VIA_SuperAccel_Module.py").exists():
            _sa_sys.path.insert(0, str(_sa_p / "supportive modules"))
            break
        _sa_p = _sa_p.parent
    import VIA_SuperAccel_Module as VIA_ACCEL  # accel_map/fetch/pip_install/run_fast
except Exception:
    VIA_ACCEL = None  # graceful:加速器缺席零影響
# ===== [VIA:ACCEL-BRIDGE:END] =====
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ACCENT = "5E7032"; HEAD_BG = "F0EEE8"; LINE = "E2DFD6"
COLS = [("ID", 6), ("時間 Time", 10), ("原始內容 Text", 64),
        ("AI分類 Category", 16), ("發話人 Speaker", 16),
        ("Action", 9), ("重點 Key", 9), ("已確認 Confirmed", 12)]


def build(db_path: str, out_dir: str, meeting_id: int = 1) -> str:
    from ..core import store
    os.makedirs(out_dir, exist_ok=True)
    proj = store.load_project(db_path)
    segs = [s for s in proj["segments"] if s["meeting_id"] == meeting_id]

    wb = Workbook(); ws = wb.active; ws.title = "Transcript_Review"
    thin = Side(style="thin", color=LINE); border = Border(bottom=thin)
    head_font = Font(name="Arial", bold=True, size=9, color="56524A")
    head_fill = PatternFill("solid", fgColor=HEAD_BG)

    for c, (name, w) in enumerate(COLS, 1):
        cell = ws.cell(1, c, name)
        cell.font = head_font; cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for r, s in enumerate(segs, 2):
        vals = [s["ord"], s["at_time"], s["text"], s.get("category", ""),
                s.get("speaker", ""), "TRUE" if s.get("action") else "FALSE",
                "TRUE" if s.get("keypt") else "FALSE",
                "TRUE" if s.get("confirmed") else "FALSE"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c == 3),
                                       horizontal="center" if c in (1, 6, 7, 8) else "left")
            if c == 4 and v:  # category colour cue
                cell.font = Font(name="Arial", size=10, bold=True, color=ACCENT)
        ws.row_dimensions[r].height = 30

    n = len(segs) + 1
    cat_dv = DataValidation(type="list", formula1='"Decision,Discussion,Action,Risk"', allow_blank=True)
    bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(cat_dv); ws.add_data_validation(bool_dv)
    cat_dv.add(f"D2:D{n}")
    for col in ("F", "G", "H"):
        bool_dv.add(f"{col}2:{col}{n}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{n}"
    # summary row using formulas (counts) — zero hardcoded calcs
    sr = n + 2
    ws.cell(sr, 2, "Action 數").font = Font(name="Arial", bold=True)
    ws.cell(sr, 3, f'=COUNTIF(F2:F{n},TRUE)')
    ws.cell(sr + 1, 2, "已確認").font = Font(name="Arial", bold=True)
    ws.cell(sr + 1, 3, f'=COUNTIF(H2:H{n},TRUE)')

    path = os.path.join(out_dir, "VRN_Meeting_Transcript.xlsx")
    wb.save(path)
    return path
