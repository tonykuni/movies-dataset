#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整PDF表格提取系統 - 包含所有庫和工具
================================================
整合所有功能的綜合解決方案，不遺漏任何庫和功能

預設路徑:
- 輸入目錄: C:/Users/tonyk/OneDrive/Desktop/VeritasReportNova/input
- 輸出目錄: C:/Users/tonyk/OneDrive/Desktop/VeritasReportNova/output
"""

import os
import sys
import json
import time
import logging
import warnings
import traceback
import subprocess
import shutil
import io
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Union
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
import multiprocessing as mp
from functools import wraps
import re
import tempfile
import base64
import hashlib
import urllib.request
import urllib.parse

# 核心數據處理
import pandas as pd
import numpy as np

# PDF處理工具 - 全面導入
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    print("⚠️ pdfplumber 未安裝: pip install pdfplumber")
    HAS_PDFPLUMBER = False

try:
    import camelot
    HAS_CAMELOT = True
except ImportError:
    print("⚠️ camelot 未安裝: pip install camelot-py[cv]")
    HAS_CAMELOT = False

try:
    import tabula
    HAS_TABULA = True
except ImportError:
    print("⚠️ tabula 未安裝: pip install tabula-py")
    HAS_TABULA = False

try:
    from pdfminer.high_level import extract_text, extract_pages
    from pdfminer.layout import LAParams, LTTextBox, LTTextLine, LTFigure
    from pdfminer.pdfpage import PDFPage
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.converter import TextConverter, PDFPageAggregator
    HAS_PDFMINER = True
except ImportError:
    print("⚠️ pdfminer 未安裝: pip install pdfminer.six")
    HAS_PDFMINER = False

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    print("⚠️ PyMuPDF 未安裝: pip install PyMuPDF")
    HAS_PYMUPDF = False

try:
    import pdf2image
    HAS_PDF2IMAGE = True
except ImportError:
    print("⚠️ pdf2image 未安裝: pip install pdf2image")
    HAS_PDF2IMAGE = False

# OCR工具
try:
    import pytesseract
    from PIL import Image, ImageEnhance, ImageFilter
    import cv2
    HAS_OCR = True
except ImportError:
    print("⚠️ OCR工具未安裝: pip install pytesseract pillow opencv-python")
    HAS_OCR = False

# Excel和文件輸出
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.formatting import Rule
    HAS_OPENPYXL = True
except ImportError:
    print("⚠️ openpyxl 未安裝: pip install openpyxl")
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    print("⚠️ xlsxwriter 未安裝: pip install xlsxwriter")
    HAS_XLSXWRITER = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    print("⚠️ reportlab 未安裝: pip install reportlab")
    HAS_REPORTLAB = False

# 網頁和JavaScript支持
try:
    import requests
    from bs4 import BeautifulSoup
    HAS_WEB = True
except ImportError:
    print("⚠️ 網頁工具未安裝: pip install requests beautifulsoup4")
    HAS_WEB = False

try:
    import selenium
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    HAS_SELENIUM = True
except ImportError:
    print("⚠️ selenium 未安裝: pip install selenium")
    HAS_SELENIUM = False

# JavaScript引擎
try:
    import js2py
    HAS_JS2PY = True
except ImportError:
    print("⚠️ js2py 未安裝: pip install js2py")
    HAS_JS2PY = False

# 系統監控
try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    print("⚠️ psutil 未安裝: pip install psutil")
    HAS_PSUTIL = False

# 數據分析和可視化
try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    HAS_PLOTTING = True
except ImportError:
    print("⚠️ 繪圖工具未安裝: pip install matplotlib seaborn")
    HAS_PLOTTING = False

# 機器學習支持
try:
    import sklearn
    from sklearn.cluster import KMeans
    from sklearn.preprocessing import StandardScaler
    HAS_SKLEARN = True
except ImportError:
    print("⚠️ sklearn 未安裝: pip install scikit-learn")
    HAS_SKLEARN = False

# 額外工具
try:
    import openpyxl_style_writer
    HAS_STYLE_WRITER = True
except ImportError:
    HAS_STYLE_WRITER = False

try:
    import pdfquery
    HAS_PDFQUERY = True
except ImportError:
    print("⚠️ pdfquery 未安裝: pip install pdfquery")
    HAS_PDFQUERY = False

# 禁用警告
warnings.filterwarnings('ignore')

# 預設路徑配置
DEFAULT_INPUT_DIR = r"C:\Users\tonyk\OneDrive\Desktop\VeritasReportNova\input"
DEFAULT_OUTPUT_DIR = r"C:\Users\tonyk\OneDrive\Desktop\VeritasReportNova\output"


class LibraryManager:
    """庫管理器 - 檢查和安裝缺失的庫"""
    
    def __init__(self):
        self.required_packages = {
            'pdfplumber': 'pdfplumber',
            'camelot': 'camelot-py[cv]',
            'tabula': 'tabula-py',
            'pdfminer': 'pdfminer.six',
            'pymupdf': 'PyMuPDF',
            'pdf2image': 'pdf2image',
            'pytesseract': 'pytesseract',
            'pillow': 'pillow',
            'opencv': 'opencv-python',
            'openpyxl': 'openpyxl',
            'xlsxwriter': 'xlsxwriter',
            'reportlab': 'reportlab',
            'requests': 'requests',
            'beautifulsoup': 'beautifulsoup4',
            'selenium': 'selenium',
            'js2py': 'js2py',
            'psutil': 'psutil',
            'matplotlib': 'matplotlib',
            'seaborn': 'seaborn',
            'sklearn': 'scikit-learn',
            'pdfquery': 'pdfquery'
        }
    
    def check_and_install_missing(self):
        """檢查並提示安裝缺失的庫"""
        missing_packages = []
        
        package_checks = {
            'pdfplumber': HAS_PDFPLUMBER,
            'camelot': HAS_CAMELOT,
            'tabula': HAS_TABULA,
            'pdfminer': HAS_PDFMINER,
            'pymupdf': HAS_PYMUPDF,
            'pdf2image': HAS_PDF2IMAGE,
            'pytesseract': HAS_OCR,
            'openpyxl': HAS_OPENPYXL,
            'xlsxwriter': HAS_XLSXWRITER,
            'reportlab': HAS_REPORTLAB,
            'requests': HAS_WEB,
            'selenium': HAS_SELENIUM,
            'js2py': HAS_JS2PY,
            'psutil': HAS_PSUTIL,
            'matplotlib': HAS_PLOTTING,
            'sklearn': HAS_SKLEARN
        }
        
        for package, is_available in package_checks.items():
            if not is_available:
                missing_packages.append(self.required_packages[package])
        
        if missing_packages:
            print("\n🚨 發現缺失的依賴包:")
            for pkg in missing_packages:
                print(f"   pip install {pkg}")
            print("\n建議執行以下命令安裝所有依賴:")
            print(f"   pip install {' '.join(missing_packages)}")
            
            install_now = input("\n是否現在安裝缺失的包? (y/n): ").strip().lower()
            if install_now == 'y':
                self._install_packages(missing_packages)
        
        return missing_packages
    
    def _install_packages(self, packages):
        """安裝缺失的包"""
        for package in packages:
            try:
                print(f"\n正在安裝 {package}...")
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"✅ {package} 安裝完成")
            except subprocess.CalledProcessError:
                print(f"❌ {package} 安裝失敗")


class JavaScriptTableExtractor:
    """JavaScript表格提取器"""
    
    def __init__(self, logger):
        self.logger = logger
        self.js_engine = None
        self._init_js_engine()
    
    def _init_js_engine(self):
        """初始化JavaScript引擎"""
        if HAS_JS2PY:
            try:
                self.js_engine = js2py
                self.logger.info("JavaScript引擎已初始化")
            except Exception as e:
                self.logger.error(f"JavaScript引擎初始化失敗: {e}")
    
    def extract_with_pdf_js(self, pdf_path: str) -> List[Dict]:
        """使用pdf.js方式提取表格"""
        if not HAS_SELENIUM:
            self.logger.warning("Selenium不可用，跳過pdf.js提取")
            return []
        
        try:
            # 這裡實現pdf.js的表格提取邏輯
            # 由於需要瀏覽器環境，這裡提供框架
            self.logger.info("正在嘗試pdf.js提取...")
            
            # 創建臨時HTML文件調用pdf.js
            html_content = self._create_pdf_js_html(pdf_path)
            
            # 使用Selenium運行
            tables = self._run_pdf_js_extraction(html_content, pdf_path)
            
            return tables
            
        except Exception as e:
            self.logger.error(f"pdf.js提取失敗: {str(e)}")
            return []
    
    def _create_pdf_js_html(self, pdf_path: str) -> str:
        """創建pdf.js HTML模板"""
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <script src="https://cdnjs.cloudflare.com/ajax/libs/pdf.js/2.16.105/pdf.min.js"></script>
        </head>
        <body>
            <canvas id="pdf-canvas"></canvas>
            <div id="extracted-tables"></div>
            
            <script>
                pdfjsLib.getDocument('{pdf_url}').promise.then(function(pdf) {{
                    // PDF.js表格提取邏輯
                    let tables = [];
                    
                    for (let pageNum = 1; pageNum <= pdf.numPages; pageNum++) {{
                        pdf.getPage(pageNum).then(function(page) {{
                            // 獲取文字內容
                            page.getTextContent().then(function(textContent) {{
                                // 分析文字結構，識別表格
                                let tableData = analyzeTextStructure(textContent);
                                if (tableData.length > 0) {{
                                    tables.push({{
                                        page: pageNum,
                                        tables: tableData
                                    }});
                                }}
                            }});
                        }});
                    }}
                    
                    // 輸出結果
                    window.extractedTables = tables;
                }});
                
                function analyzeTextStructure(textContent) {{
                    // 簡化的表格識別邏輯
                    let items = textContent.items;
                    let lines = [];
                    let currentLine = [];
                    let lastY = null;
                    
                    items.forEach(function(item) {{
                        if (lastY !== null && Math.abs(item.transform[5] - lastY) > 5) {{
                            if (currentLine.length > 0) {{
                                lines.push(currentLine);
                                currentLine = [];
                            }}
                        }}
                        currentLine.push(item.str);
                        lastY = item.transform[5];
                    }});
                    
                    if (currentLine.length > 0) {{
                        lines.push(currentLine);
                    }}
                    
                    // 檢測表格結構
                    return detectTables(lines);
                }}
                
                function detectTables(lines) {{
                    let tables = [];
                    let currentTable = [];
                    
                    lines.forEach(function(line) {{
                        if (line.length > 2) {{ // 假設表格至少有3列
                            currentTable.push(line);
                        }} else if (currentTable.length > 2) {{ // 至少3行
                            tables.push(currentTable);
                            currentTable = [];
                        }}
                    }});
                    
                    if (currentTable.length > 2) {{
                        tables.push(currentTable);
                    }}
                    
                    return tables;
                }}
            </script>
        </body>
        </html>
        """
        
        return html_template.format(pdf_url=f"file:///{pdf_path.replace('\\', '/')}")
    
    def _run_pdf_js_extraction(self, html_content: str, pdf_path: str) -> List[Dict]:
        """運行pdf.js提取"""
        try:
            # 創建臨時HTML文件
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
                f.write(html_content)
                html_file = f.name
            
            # 使用Selenium運行
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            
            driver = webdriver.Chrome(options=chrome_options)
            driver.get(f"file://{html_file}")
            
            # 等待處理完成
            time.sleep(5)
            
            # 獲取結果
            tables = driver.execute_script("return window.extractedTables || [];")
            
            driver.quit()
            os.unlink(html_file)
            
            # 轉換為標準格式
            result_tables = []
            for page_data in tables:
                for i, table_data in enumerate(page_data.get('tables', [])):
                    df = pd.DataFrame(table_data)
                    result_tables.append({
                        'page': page_data['page'],
                        'table_num': i + 1,
                        'method': 'pdf.js',
                        'data': df,
                        'shape': df.shape,
                        'quality_score': self._calculate_quality_score(df)
                    })
            
            return result_tables
            
        except Exception as e:
            self.logger.error(f"pdf.js運行失敗: {str(e)}")
            return []
    
    def extract_with_pdf2table(self, pdf_path: str) -> List[Dict]:
        """使用pdf2table方式提取"""
        if not self.js_engine:
            self.logger.warning("JavaScript引擎不可用")
            return []
        
        try:
            # 模擬pdf2table的JavaScript邏輯
            js_code = """
            function extractTables(pdfData) {
                // 這裡實現pdf2table的邏輯
                var tables = [];
                
                // 簡化的表格檢測邏輯
                if (pdfData && pdfData.length > 0) {
                    // 假設我們有PDF的文字數據
                    var lines = pdfData.split('\\n');
                    var currentTable = [];
                    
                    for (var i = 0; i < lines.length; i++) {
                        var line = lines[i].trim();
                        if (line.length > 0) {
                            var cells = line.split(/\\s{2,}|\\t/); // 分割多個空格或tab
                            if (cells.length > 1) {
                                currentTable.push(cells);
                            } else if (currentTable.length > 2) {
                                tables.push(currentTable);
                                currentTable = [];
                            }
                        }
                    }
                    
                    if (currentTable.length > 2) {
                        tables.push(currentTable);
                    }
                }
                
                return tables;
            }
            """
            
            # 執行JavaScript代碼
            extract_func = self.js_engine.eval_js(js_code)
            
            # 獲取PDF文字內容
            if HAS_PDFMINER:
                pdf_text = extract_text(pdf_path)
                js_tables = extract_func(pdf_text)
                
                # 轉換為DataFrame格式
                result_tables = []
                for i, table_data in enumerate(js_tables):
                    if table_data:
                        df = pd.DataFrame(table_data[1:], columns=table_data[0] if table_data else [])
                        result_tables.append({
                            'page': 1,
                            'table_num': i + 1,
                            'method': 'pdf2table',
                            'data': df,
                            'shape': df.shape,
                            'quality_score': self._calculate_quality_score(df)
                        })
                
                return result_tables
            else:
                self.logger.warning("需要pdfminer來獲取PDF文字內容")
                return []
                
        except Exception as e:
            self.logger.error(f"pdf2table提取失敗: {str(e)}")
            return []
    
    def _calculate_quality_score(self, df: pd.DataFrame) -> float:
        """計算質量分數"""
        if df.empty:
            return 0.0
        
        rows, cols = df.shape
        if rows == 0 or cols == 0:
            return 0.0
        
        # 計算非空單元格比例
        non_empty = df.count().sum()
        total_cells = rows * cols
        fill_rate = non_empty / total_cells if total_cells > 0 else 0
        
        return round(fill_rate * 0.8 + (min(rows, cols) / max(rows, cols)) * 0.2, 3)


class EnhancedTableExtractor:
    """增強版表格提取器 - 包含所有工具"""
    
    def __init__(self, logger):
        self.logger = logger
        self.js_extractor = JavaScriptTableExtractor(logger)
        self.available_tools = self._check_all_tools()
    
    def _check_all_tools(self) -> Dict[str, bool]:
        """檢查所有工具的可用性"""
        return {
            'pdfplumber': HAS_PDFPLUMBER,
            'camelot': HAS_CAMELOT,
            'tabula': HAS_TABULA,
            'pdfminer': HAS_PDFMINER,
            'pymupdf': HAS_PYMUPDF,
            'pdf2image': HAS_PDF2IMAGE,
            'ocr': HAS_OCR,
            'openpyxl': HAS_OPENPYXL,
            'xlsxwriter': HAS_XLSXWRITER,
            'reportlab': HAS_REPORTLAB,
            'web': HAS_WEB,
            'selenium': HAS_SELENIUM,
            'js2py': HAS_JS2PY,
            'psutil': HAS_PSUTIL,
            'plotting': HAS_PLOTTING,
            'sklearn': HAS_SKLEARN,
            'pdfquery': HAS_PDFQUERY
        }
    
    def extract_all_methods(self, pdf_path: str) -> Dict[str, List[Dict]]:
        """使用所有可用方法提取表格"""
        self.logger.info(f"開始全方位提取: {pdf_path}")
        
        all_results = {}
        
        # 1. pdfplumber 提取
        if self.available_tools['pdfplumber']:
            try:
                tables = self._extract_with_pdfplumber(pdf_path)
                if tables:
                    all_results['pdfplumber'] = tables
                    self.logger.info(f"pdfplumber: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"pdfplumber失敗: {e}")
        
        # 2. camelot 提取  
        if self.available_tools['camelot']:
            try:
                tables = self._extract_with_camelot(pdf_path)
                if tables:
                    all_results['camelot'] = tables
                    self.logger.info(f"camelot: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"camelot失敗: {e}")
        
        # 3. tabula 提取
        if self.available_tools['tabula']:
            try:
                tables = self._extract_with_tabula(pdf_path)
                if tables:
                    all_results['tabula'] = tables
                    self.logger.info(f"tabula: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"tabula失敗: {e}")
        
        # 4. pdfminer 提取
        if self.available_tools['pdfminer']:
            try:
                tables = self._extract_with_pdfminer(pdf_path)
                if tables:
                    all_results['pdfminer'] = tables
                    self.logger.info(f"pdfminer: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"pdfminer失敗: {e}")
        
        # 5. PyMuPDF 提取
        if self.available_tools['pymupdf']:
            try:
                tables = self._extract_with_pymupdf(pdf_path)
                if tables:
                    all_results['pymupdf'] = tables
                    self.logger.info(f"pymupdf: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"pymupdf失敗: {e}")
        
        # 6. OCR 提取
        if self.available_tools['ocr']:
            try:
                tables = self._extract_with_enhanced_ocr(pdf_path)
                if tables:
                    all_results['ocr'] = tables
                    self.logger.info(f"OCR: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"OCR失敗: {e}")
        
        # 7. JavaScript工具提取
        try:
            js_tables = self.js_extractor.extract_with_pdf_js(pdf_path)
            if js_tables:
                all_results['pdf.js'] = js_tables
                self.logger.info(f"pdf.js: {len(js_tables)} 個表格")
        except Exception as e:
            self.logger.error(f"pdf.js失敗: {e}")
        
        try:
            pdf2table_results = self.js_extractor.extract_with_pdf2table(pdf_path)
            if pdf2table_results:
                all_results['pdf2table'] = pdf2table_results
                self.logger.info(f"pdf2table: {len(pdf2table_results)} 個表格")
        except Exception as e:
            self.logger.error(f"pdf2table失敗: {e}")
        
        # 8. pdfquery 提取
        if self.available_tools['pdfquery']:
            try:
                tables = self._extract_with_pdfquery(pdf_path)
                if tables:
                    all_results['pdfquery'] = tables
                    self.logger.info(f"pdfquery: {len(tables)} 個表格")
            except Exception as e:
                self.logger.error(f"pdfquery失敗: {e}")
        
        return all_results
    
    def _extract_with_pdfplumber(self, pdf_path: str) -> List[Dict]:
        """使用pdfplumber提取表格"""
        tables = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_tables = page.extract_tables()
                    for table_num, table in enumerate(page_tables, 1):
                        if table and len(table) > 0:
                            headers = table[0] if table[0] else [f"Column_{i}" for i in range(len(table[0]) if table else 0)]
                            data = table[1:] if len(table) > 1 else []
                            
                            df = pd.DataFrame(data, columns=headers)
                            
                            tables.append({
                                'page': page_num,
                                'table_num': table_num,
                                'method': 'pdfplumber',
                                'data': df,
                                'raw_data': table,
                                'shape': df.shape,
                                'quality_score': self._calculate_quality_score(df)
                            })
            
            return tables
            
        except Exception as e:
            self.logger.error(f"pdfplumber提取失敗: {str(e)}")
            return []
    
    def _extract_with_camelot(self, pdf_path: str) -> List[Dict]:
        """使用camelot提取表格"""
        tables = []
        try:
            # 嘗試兩種flavor
            for flavor in ['lattice', 'stream']:
                try:
                    camelot_tables = camelot.read_pdf(str(pdf_path), pages='all', flavor=flavor)
                    
                    for i, table in enumerate(camelot_tables, 1):
                        df = table.df
                        tables.append({
                            'page': table.page,
                            'table_num': i,
                            'method': f'camelot-{flavor}',
                            'data': df,
                            'shape': df.shape,
                            'accuracy': getattr(table, 'accuracy', 0),
                            'quality_score': self._calculate_quality_score(df)
                        })
                    
                    if tables:  # 如果找到表格就不嘗試其他flavor
                        break
                        
                except Exception as e:
                    self.logger.debug(f"camelot {flavor} 失敗: {e}")
                    continue
            
            return tables
            
        except Exception as e:
            self.logger.error(f"camelot提取失敗: {str(e)}")
            return []
    
    def _extract_with_tabula(self, pdf_path: str) -> List[Dict]:
        """使用tabula提取表格"""
        tables = []
        try:
            # 嘗試不同的參數組合
            param_combinations = [
                {'pages': 'all', 'multiple_tables': True},
                {'pages': 'all', 'multiple_tables': True, 'stream': True},
                {'pages': 'all', 'multiple_tables': True, 'lattice': True}
            ]
            
            for params in param_combinations:
                try:
                    tabula_tables = tabula.read_pdf(str(pdf_path), **params)
                    
                    for i, df in enumerate(tabula_tables, 1):
                        if not df.empty:
                            tables.append({
                                'page': 1,  # tabula不提供精確頁碼
                                'table_num': i,
                                'method': f'tabula-{list(params.keys())}',
                                'data': df,
                                'shape': df.shape,
                                'quality_score': self._calculate_quality_score(df)
                            })
                    
                    if tables:  # 找到表格就停止
                        break
                        
                except Exception as e:
                    self.logger.debug(f"tabula參數 {params} 失敗: {e}")
                    continue
            
            return tables
            
        except Exception as e:
            self.logger.error(f"tabula提取失敗: {str(e)}")
            return []
    
    def _extract_with_pdfminer(self, pdf_path: str) -> List[Dict]:
        """使用pdfminer提取表格"""
        tables = []
        try:
            # 提取文字內容並分析結構
            text_content = extract_text(pdf_path)
            
            # 簡單的表格檢測邏輯
            lines = text_content.split('\n')
            potential_tables = []
            current_table = []
            
            for line in lines:
                line = line.strip()
                if not line:
                    if len(current_table) > 2:  # 至少3行
                        potential_tables.append(current_table)
                    current_table = []
                    continue
                
                # 檢測是否為表格行（包含多個用空格或tab分隔的項目）
                cells = re.split(r'\s{2,}|\t+', line)
                if len(cells) > 1:
                    current_table.append(cells)
                elif len(current_table) > 2:
                    potential_tables.append(current_table)
                    current_table = []
            
            if len(current_table) > 2:
                potential_tables.append(current_table)
            
            # 轉換為DataFrame
            for i, table_data in enumerate(potential_tables, 1):
                if table_data:
                    try:
                        headers = table_data[0]
                        data = table_data[1:]
                        df = pd.DataFrame(data, columns=headers)
                        
                        tables.append({
                            'page': 1,
                            'table_num': i,
                            'method': 'pdfminer',
                            'data': df,
                            'shape': df.shape,
                            'quality_score': self._calculate_quality_score(df)
                        })
                    except Exception as e:
                        self.logger.debug(f"pdfminer表格 {i} 轉換失敗: {e}")
            
            return tables
            
        except Exception as e:
            self.logger.error(f"pdfminer提取失敗: {str(e)}")
            return []
    
    def _extract_with_pymupdf(self, pdf_path: str) -> List[Dict]:
        """使用PyMuPDF提取表格"""
        tables = []
        try:
            pdf_document = fitz.open(pdf_path)
            
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                
                # 嘗試查找表格
                page_tables = page.find_tables()
                
                for table_num, table in enumerate(page_tables, 1):
                    try:
                        # 提取表格數據
                        table_data = table.extract()
                        if table_data and len(table_data) > 1:
                            headers = table_data[0]
                            data = table_data[1:]
                            df = pd.DataFrame(data, columns=headers)
                            
                            tables.append({
                                'page': page_num + 1,
                                'table_num': table_num,
                                'method': 'pymupdf',
                                'data': df,
                                'shape': df.shape,
                                'quality_score': self._calculate_quality_score(df)
                            })
                    except Exception as e:
                        self.logger.debug(f"PyMuPDF表格 {table_num} 提取失敗: {e}")
            
            pdf_document.close()
            return tables
            
        except Exception as e:
            self.logger.error(f"PyMuPDF提取失敗: {str(e)}")
            return []
    
    def _extract_with_enhanced_ocr(self, pdf_path: str) -> List[Dict]:
        """使用增強版OCR提取表格"""
        if not (HAS_OCR and HAS_PDF2IMAGE):
            return []
        
        tables = []
        try:
            # 將PDF轉為圖像
            images = pdf2image.convert_from_path(pdf_path, dpi=300)
            
            for page_num, image in enumerate(images, 1):
                # 圖像預處理
                enhanced_image = self._enhance_image_for_ocr(image)
                
                # OCR識別
                ocr_data = pytesseract.image_to_data(enhanced_image, output_type=pytesseract.Output.DICT, lang='chi_tra+eng')
                
                # 分析OCR結果並檢測表格
                page_tables = self._detect_tables_from_ocr(ocr_data)
                
                for table_num, table_data in enumerate(page_tables, 1):
                    if table_data:
                        df = pd.DataFrame(table_data)
                        tables.append({
                            'page': page_num,
                            'table_num': table_num,
                            'method': 'enhanced-ocr',
                            'data': df,
                            'shape': df.shape,
                            'quality_score': self._calculate_quality_score(df)
                        })
            
            return tables
            
        except Exception as e:
            self.logger.error(f"增強OCR提取失敗: {str(e)}")
            return []
    
    def _extract_with_pdfquery(self, pdf_path: str) -> List[Dict]:
        """使用pdfquery提取表格"""
        tables = []
        try:
            import pdfquery
            
            pdf = pdfquery.PDFQuery(pdf_path)
            pdf.load()
            
            # 查找可能的表格結構
            # 這是一個簡化的實現
            elements = pdf.tree.iter()
            
            # 簡單的表格檢測邏輯
            # 實際實現會更複雜
            table_candidates = []
            for elem in elements:
                if elem.tag == 'LTTextBox':
                    # 分析文字框的內容
                    text = elem.text if hasattr(elem, 'text') else ''
                    if self._looks_like_table(text):
                        table_candidates.append(text)
            
            # 處理候選表格
            for i, table_text in enumerate(table_candidates, 1):
                df = self._parse_table_text(table_text)
                if not df.empty:
                    tables.append({
                        'page': 1,
                        'table_num': i,
                        'method': 'pdfquery',
                        'data': df,
                        'shape': df.shape,
                        'quality_score': self._calculate_quality_score(df)
                    })
            
            return tables
            
        except Exception as e:
            self.logger.error(f"pdfquery提取失敗: {str(e)}")
            return []
    
    def _enhance_image_for_ocr(self, image):
        """增強圖像以提高OCR效果"""
        try:
            # 轉換為OpenCV格式
            img_array = np.array(image)
            img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
            
            # 灰度轉換
            gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
            
            # 降噪
            denoised = cv2.fastNlMeansDenoising(gray)
            
            # 增強對比度
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(denoised)
            
            # 二值化
            _, binary = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # 轉回PIL格式
            return Image.fromarray(binary)
            
        except Exception as e:
            self.logger.debug(f"圖像增強失敗: {e}")
            return image
    
    def _detect_tables_from_ocr(self, ocr_data) -> List[List[List[str]]]:
        """從OCR數據檢測表格"""
        try:
            # 根據位置信息組織文字
            words = []
            for i in range(len(ocr_data['text'])):
                if int(ocr_data['conf'][i]) > 30:  # 置信度閾值
                    words.append({
                        'text': ocr_data['text'][i],
                        'left': ocr_data['left'][i],
                        'top': ocr_data['top'][i],
                        'width': ocr_data['width'][i],
                        'height': ocr_data['height'][i]
                    })
            
            # 按行分組
            rows = self._group_words_by_rows(words)
            
            # 檢測表格
            tables = []
            current_table = []
            
            for row in rows:
                if len(row) > 1:  # 至少2個單元格
                    current_table.append([word['text'] for word in row])
                elif len(current_table) > 2:  # 至少3行
                    tables.append(current_table)
                    current_table = []
            
            if len(current_table) > 2:
                tables.append(current_table)
            
            return tables
            
        except Exception as e:
            self.logger.debug(f"OCR表格檢測失敗: {e}")
            return []
    
    def _group_words_by_rows(self, words) -> List[List[Dict]]:
        """按行分組單詞"""
        if not words:
            return []
        
        # 按y坐標排序
        words.sort(key=lambda w: w['top'])
        
        rows = []
        current_row = [words[0]]
        row_top = words[0]['top']
        row_tolerance = 10  # 像素容差
        
        for word in words[1:]:
            if abs(word['top'] - row_top) <= row_tolerance:
                current_row.append(word)
            else:
                # 按x坐標排序當前行
                current_row.sort(key=lambda w: w['left'])
                rows.append(current_row)
                
                current_row = [word]
                row_top = word['top']
        
        if current_row:
            current_row.sort(key=lambda w: w['left'])
            rows.append(current_row)
        
        return rows
    
    def _looks_like_table(self, text: str) -> bool:
        """判斷文字是否像表格"""
        lines = text.strip().split('\n')
        if len(lines) < 3:
            return False
        
        # 檢查是否有規律的分隔符
        for line in lines[:3]:
            if len(re.split(r'\s{2,}|\t+', line)) > 1:
                return True
        
        return False
    
    def _parse_table_text(self, text: str) -> pd.DataFrame:
        """解析表格文字"""
        try:
            lines = text.strip().split('\n')
            table_data = []
            
            for line in lines:
                cells = re.split(r'\s{2,}|\t+', line.strip())
                if len(cells) > 1:
                    table_data.append(cells)
            
            if len(table_data) > 1:
                return pd.DataFrame(table_data[1:], columns=table_data[0])
            else:
                return pd.DataFrame()
                
        except Exception:
            return pd.DataFrame()
    
    def _calculate_quality_score(self, df: pd.DataFrame) -> float:
        """計算質量分數"""
        if df.empty:
            return 0.0
        
        try:
            rows, cols = df.shape
            if rows == 0 or cols == 0:
                return 0.0
            
            # 計算非空單元格比例
            non_empty = df.count().sum()
            total_cells = rows * cols
            fill_rate = non_empty / total_cells if total_cells > 0 else 0
            
            # 計算數據類型多樣性
            numeric_cols = len(df.select_dtypes(include=[np.number]).columns)
            type_diversity = min(1.0, (numeric_cols + (cols - numeric_cols)) / cols)
            
            # 綜合質量分數
            quality_score = (fill_rate * 0.7 + type_diversity * 0.3)
            return round(quality_score, 3)
            
        except Exception:
            return 0.0


class AdvancedAnalyzer:
    """高級分析器 - 提供深度分析功能"""
    
    def __init__(self, logger):
        self.logger = logger
    
    def analyze_all_results(self, all_results: Dict[str, Dict[str, List[Dict]]]) -> Dict[str, Any]:
        """分析所有提取結果"""
        self.logger.info("開始深度分析...")
        
        analysis = {
            'summary': {},
            'method_comparison': {},
            'quality_analysis': {},
            'consensus_tables': [],
            'recommendations': []
        }
        
        # 統計摘要
        total_files = len(all_results)
        total_methods = 0
        total_tables = 0
        
        method_stats = {}
        
        for file_name, file_results in all_results.items():
            for method, tables in file_results.items():
                if method not in method_stats:
                    method_stats[method] = {
                        'files': 0,
                        'tables': 0,
                        'avg_quality': 0,
                        'total_quality': 0
                    }
                
                method_stats[method]['files'] += 1
                method_stats[method]['tables'] += len(tables)
                
                for table in tables:
                    method_stats[method]['total_quality'] += table['quality_score']
                    total_tables += 1
        
        # 計算平均質量
        for method, stats in method_stats.items():
            if stats['tables'] > 0:
                stats['avg_quality'] = stats['total_quality'] / stats['tables']
        
        analysis['summary'] = {
            'total_files': total_files,
            'total_methods': len(method_stats),
            'total_tables': total_tables,
            'avg_tables_per_file': total_tables / total_files if total_files > 0 else 0
        }
        
        analysis['method_comparison'] = method_stats
        
        # 質量分析
        all_quality_scores = []
        for file_results in all_results.values():
            for tables in file_results.values():
                for table in tables:
                    all_quality_scores.append(table['quality_score'])
        
        if all_quality_scores:
            analysis['quality_analysis'] = {
                'mean_quality': np.mean(all_quality_scores),
                'std_quality': np.std(all_quality_scores),
                'min_quality': np.min(all_quality_scores),
                'max_quality': np.max(all_quality_scores),
                'high_quality_count': len([q for q in all_quality_scores if q > 0.7]),
                'medium_quality_count': len([q for q in all_quality_scores if 0.4 <= q <= 0.7]),
                'low_quality_count': len([q for q in all_quality_scores if q < 0.4])
            }
        
        # 生成建議
        analysis['recommendations'] = self._generate_recommendations(analysis)
        
        return analysis
    
    def _generate_recommendations(self, analysis: Dict[str, Any]) -> List[str]:
        """生成改進建議"""
        recommendations = []
        
        # 基於方法比較的建議
        method_comparison = analysis.get('method_comparison', {})
        if method_comparison:
            best_method = max(method_comparison.keys(), 
                            key=lambda m: method_comparison[m]['avg_quality'])
            recommendations.append(f"推薦主要使用 {best_method} 方法，平均質量最高")
        
        # 基於質量分析的建議
        quality_analysis = analysis.get('quality_analysis', {})
        if quality_analysis:
            if quality_analysis['low_quality_count'] > quality_analysis['high_quality_count']:
                recommendations.append("檢測到較多低質量表格，建議檢查PDF文件品質或嘗試OCR")
            
            if quality_analysis['std_quality'] > 0.3:
                recommendations.append("質量分數變異較大，建議標準化PDF格式")
        
        recommendations.append("定期更新提取工具以獲得最佳效果")
        recommendations.append("考慮建立標準化的表格格式模板")
        
        return recommendations


class ComprehensiveExporter:
    """綜合導出器 - 支持所有格式"""
    
    def __init__(self, logger, output_dir: Path):
        self.logger = logger
        self.output_dir = output_dir
        self.available_tools = {
            'openpyxl': HAS_OPENPYXL,
            'xlsxwriter': HAS_XLSXWRITER,
            'reportlab': HAS_REPORTLAB,
            'plotting': HAS_PLOTTING
        }
    
    def export_all_formats(self, all_results: Dict[str, Dict[str, List[Dict]]], 
                          analysis: Dict[str, Any]) -> Dict[str, str]:
        """導出所有格式"""
        self.logger.info("開始導出所有格式...")
        
        export_paths = {}
        
        # 1. Excel格式 (多種方式)
        if self.available_tools['openpyxl']:
            excel_path = self._export_to_excel_openpyxl(all_results, analysis)
            if excel_path:
                export_paths['excel_openpyxl'] = excel_path
        
        if self.available_tools['xlsxwriter']:
            excel_xlsxwriter_path = self._export_to_excel_xlsxwriter(all_results, analysis)
            if excel_xlsxwriter_path:
                export_paths['excel_xlsxwriter'] = excel_xlsxwriter_path
        
        # 2. CSV格式
        csv_paths = self._export_to_csv(all_results)
        if csv_paths:
            export_paths['csv'] = csv_paths
        
        # 3. JSON格式
        json_path = self._export_to_json(all_results, analysis)
        if json_path:
            export_paths['json'] = json_path
        
        # 4. XML格式
        xml_path = self._export_to_xml(all_results, analysis)
        if xml_path:
            export_paths['xml'] = xml_path
        
        # 5. HTML報告
        html_path = self._export_to_html(all_results, analysis)
        if html_path:
            export_paths['html'] = html_path
        
        # 6. PDF報告
        if self.available_tools['reportlab']:
            pdf_path = self._export_to_pdf_report(all_results, analysis)
            if pdf_path:
                export_paths['pdf'] = pdf_path
        
        # 7. TXT報告
        txt_path = self._export_to_txt(all_results, analysis)
        if txt_path:
            export_paths['txt'] = txt_path
        
        # 8. 可視化圖表
        if self.available_tools['plotting']:
            chart_paths = self._export_charts(analysis)
            if chart_paths:
                export_paths['charts'] = chart_paths
        
        return export_paths
    
    def _export_to_excel_openpyxl(self, all_results: Dict, analysis: Dict) -> str:
        """使用openpyxl導出Excel"""
        try:
            output_path = self.output_dir / "comprehensive_results_openpyxl.xlsx"
            wb = Workbook()
            wb.remove(wb.active)  # 移除默認工作表
            
            # 創建摘要工作表
            summary_ws = wb.create_sheet("分析摘要", 0)
            self._create_summary_sheet_openpyxl(summary_ws, analysis)
            
            # 為每個文件和方法創建工作表
            sheet_index = 1
            for file_name, file_results in all_results.items():
                for method, tables in file_results.items():
                    if tables:
                        sheet_name = f"{self._clean_sheet_name(file_name)[:15]}_{method[:10]}"
                        ws = wb.create_sheet(sheet_name, sheet_index)
                        self._fill_method_sheet_openpyxl(ws, tables, file_name, method)
                        sheet_index += 1
            
            wb.save(output_path)
            self.logger.info(f"Excel (openpyxl) 已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Excel (openpyxl) 導出失敗: {e}")
            return ""
    
    def _export_to_excel_xlsxwriter(self, all_results: Dict, analysis: Dict) -> str:
        """使用xlsxwriter導出Excel"""
        try:
            output_path = self.output_dir / "comprehensive_results_xlsxwriter.xlsx"
            workbook = xlsxwriter.Workbook(str(output_path))
            
            # 創建格式
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 14,
                'bg_color': '#366092',
                'font_color': 'white'
            })
            
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9E2F3',
                'border': 1
            })
            
            # 創建摘要工作表
            summary_ws = workbook.add_worksheet("分析摘要")
            self._create_summary_sheet_xlsxwriter(summary_ws, analysis, title_format, header_format)
            
            # 為每個文件和方法創建工作表
            for file_name, file_results in all_results.items():
                for method, tables in file_results.items():
                    if tables:
                        sheet_name = f"{self._clean_sheet_name(file_name)[:15]}_{method[:10]}"
                        ws = workbook.add_worksheet(sheet_name)
                        self._fill_method_sheet_xlsxwriter(ws, tables, file_name, method, header_format)
            
            workbook.close()
            self.logger.info(f"Excel (xlsxwriter) 已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Excel (xlsxwriter) 導出失敗: {e}")
            return ""
    
    def _export_to_csv(self, all_results: Dict) -> List[str]:
        """導出CSV文件"""
        try:
            csv_paths = []
            csv_dir = self.output_dir / "csv_tables"
            csv_dir.mkdir(exist_ok=True)
            
            for file_name, file_results in all_results.items():
                for method, tables in file_results.items():
                    for i, table in enumerate(tables, 1):
                        csv_filename = f"{self._clean_filename(file_name)}_{method}_table_{i}.csv"
                        csv_path = csv_dir / csv_filename
                        
                        table['data'].to_csv(csv_path, index=False, encoding='utf-8-sig')
                        csv_paths.append(str(csv_path))
            
            self.logger.info(f"已導出 {len(csv_paths)} 個CSV文件")
            return csv_paths
            
        except Exception as e:
            self.logger.error(f"CSV導出失敗: {e}")
            return []
    
    def _export_to_json(self, all_results: Dict, analysis: Dict) -> str:
        """導出JSON文件"""
        try:
            output_path = self.output_dir / "comprehensive_results.json"
            
            # 序列化數據
            json_data = {
                'analysis': analysis,
                'extraction_time': datetime.now().isoformat(),
                'files': {}
            }
            
            for file_name, file_results in all_results.items():
                json_data['files'][file_name] = {}
                for method, tables in file_results.items():
                    json_data['files'][file_name][method] = []
                    for table in tables:
                        table_json = {
                            'page': table['page'],
                            'table_num': table['table_num'],
                            'method': table['method'],
                            'shape': table['shape'],
                            'quality_score': table['quality_score'],
                            'data': table['data'].to_dict('records')
                        }
                        json_data['files'][file_name][method].append(table_json)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"JSON已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"JSON導出失敗: {e}")
            return ""
    
    def _export_to_xml(self, all_results: Dict, analysis: Dict) -> str:
        """導出XML文件"""
        try:
            output_path = self.output_dir / "comprehensive_results.xml"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
                f.write('<pdf_table_extraction>\n')
                f.write(f'  <extraction_time>{datetime.now().isoformat()}</extraction_time>\n')
                
                # 寫入分析摘要
                f.write('  <analysis>\n')
                if 'summary' in analysis:
                    f.write('    <summary>\n')
                    for key, value in analysis['summary'].items():
                        f.write(f'      <{key}>{value}</{key}>\n')
                    f.write('    </summary>\n')
                f.write('  </analysis>\n')
                
                # 寫入提取結果
                f.write('  <extraction_results>\n')
                for file_name, file_results in all_results.items():
                    f.write(f'    <file name="{file_name}">\n')
                    for method, tables in file_results.items():
                        f.write(f'      <method name="{method}">\n')
                        for table in tables:
                            f.write(f'        <table page="{table["page"]}" num="{table["table_num"]}" quality="{table["quality_score"]}">\n')
                            f.write(f'          <shape rows="{table["shape"][0]}" cols="{table["shape"][1]}" />\n')
                            f.write('          <data>\n')
                            
                            df = table['data']
                            for _, row in df.iterrows():
                                f.write('            <row>\n')
                                for col, value in row.items():
                                    f.write(f'              <cell column="{col}">{str(value)}</cell>\n')
                                f.write('            </row>\n')
                            
                            f.write('          </data>\n')
                            f.write('        </table>\n')
                        f.write('      </method>\n')
                    f.write('    </file>\n')
                f.write('  </extraction_results>\n')
                f.write('</pdf_table_extraction>\n')
            
            self.logger.info(f"XML已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"XML導出失敗: {e}")
            return ""
    
    def _export_to_html(self, all_results: Dict, analysis: Dict) -> str:
        """導出HTML報告"""
        try:
            output_path = self.output_dir / "comprehensive_report.html"
            
            html_content = self._generate_comprehensive_html(all_results, analysis)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"HTML報告已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"HTML導出失敗: {e}")
            return ""
    
    def _generate_comprehensive_html(self, all_results: Dict, analysis: Dict) -> str:
        """生成綜合HTML報告"""
        html = f"""
        <!DOCTYPE html>
        <html lang="zh-TW">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>綜合PDF表格提取報告</title>
            <style>
                body {{
                    font-family: 'Microsoft JhengHei', Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                }}
                .container {{
                    max-width: 1400px;
                    margin: 0 auto;
                    background: white;
                    padding: 30px;
                    border-radius: 15px;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                }}
                h1 {{
                    color: #2c3e50;
                    text-align: center;
                    font-size: 2.5rem;
                    margin-bottom: 20px;
                    text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
                }}
                h2 {{
                    color: #34495e;
                    border-left: 4px solid #3498db;
                    padding-left: 15px;
                    margin-top: 30px;
                }}
                .summary-grid {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                    gap: 20px;
                    margin: 20px 0;
                }}
                .summary-card {{
                    background: linear-gradient(135deg, #3498db, #2980b9);
                    color: white;
                    padding: 20px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                }}
                .summary-card h3 {{
                    margin: 0 0 10px 0;
                    font-size: 1.2rem;
                }}
                .summary-card .value {{
                    font-size: 2rem;
                    font-weight: bold;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                    background: white;
                    border-radius: 8px;
                    overflow: hidden;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 12px;
                    text-align: left;
                }}
                th {{
                    background: linear-gradient(135deg, #34495e, #2c3e50);
                    color: white;
                    font-weight: bold;
                }}
                tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                tr:hover {{
                    background-color: #e8f4f8;
                }}
                .method-section {{
                    margin: 30px 0;
                    padding: 20px;
                    border: 1px solid #ddd;
                    border-radius: 10px;
                    background: #f8f9fa;
                }}
                .quality-high {{
                    background-color: #27ae60;
                    color: white;
                    padding: 4px 8px;
                    border-radius: 4px;
                }}
                .quality-medium {{
                    background-color: #f39c12;
                    color: white;
                    padding: 4px 8px;
                    border-radius: 4px;
                }}
                .quality-low {{
                    background-color: #e74c3c;
                    color: white;
                    padding: 4px 8px;
                    border-radius: 4px;
                }}
                .recommendations {{
                    background: #ecf0f1;
                    padding: 20px;
                    border-radius: 10px;
                    margin: 20px 0;
                }}
                .recommendations ul {{
                    list-style-type: none;
                    padding: 0;
                }}
                .recommendations li {{
                    background: white;
                    margin: 10px 0;
                    padding: 15px;
                    border-left: 4px solid #3498db;
                    border-radius: 5px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🚀 綜合PDF表格提取分析報告</h1>
                <p style="text-align: center; color: #7f8c8d; font-size: 1.1rem;">
                    生成時間: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}
                </p>
                
                {self._generate_summary_section(analysis)}
                {self._generate_method_comparison_section(analysis)}
                {self._generate_quality_analysis_section(analysis)}
                {self._generate_detailed_results_section(all_results)}
                {self._generate_recommendations_section(analysis)}
            </div>
        </body>
        </html>
        """
        return html
    
    def _generate_summary_section(self, analysis: Dict) -> str:
        """生成摘要部分"""
        if 'summary' not in analysis:
            return ""
        
        summary = analysis['summary']
        return f"""
        <h2>📊 提取摘要</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <h3>處理文件數</h3>
                <div class="value">{summary.get('total_files', 0)}</div>
            </div>
            <div class="summary-card">
                <h3>使用方法數</h3>
                <div class="value">{summary.get('total_methods', 0)}</div>
            </div>
            <div class="summary-card">
                <h3>提取表格總數</h3>
                <div class="value">{summary.get('total_tables', 0)}</div>
            </div>
            <div class="summary-card">
                <h3>平均表格/文件</h3>
                <div class="value">{summary.get('avg_tables_per_file', 0):.1f}</div>
            </div>
        </div>
        """
    
    def _generate_method_comparison_section(self, analysis: Dict) -> str:
        """生成方法比較部分"""
        if 'method_comparison' not in analysis:
            return ""
        
        methods = analysis['method_comparison']
        html = """
        <h2>🔍 方法比較分析</h2>
        <table>
            <thead>
                <tr>
                    <th>提取方法</th>
                    <th>處理文件數</th>
                    <th>提取表格數</th>
                    <th>平均質量分數</th>
                    <th>效能評級</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for method, stats in methods.items():
            quality = stats['avg_quality']
            if quality > 0.7:
                quality_class = "quality-high"
                rating = "優秀"
            elif quality > 0.4:
                quality_class = "quality-medium" 
                rating = "良好"
            else:
                quality_class = "quality-low"
                rating = "需改進"
            
            html += f"""
                <tr>
                    <td><strong>{method}</strong></td>
                    <td>{stats['files']}</td>
                    <td>{stats['tables']}</td>
                    <td>{quality:.3f}</td>
                    <td><span class="{quality_class}">{rating}</span></td>
                </tr>
            """
        
        html += """
            </tbody>
        </table>
        """
        
        return html
    
    def _generate_quality_analysis_section(self, analysis: Dict) -> str:
        """生成質量分析部分"""
        if 'quality_analysis' not in analysis:
            return ""
        
        qa = analysis['quality_analysis']
        return f"""
        <h2>⭐ 質量分析</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <h3>平均質量</h3>
                <div class="value">{qa.get('mean_quality', 0):.3f}</div>
            </div>
            <div class="summary-card">
                <h3>最高質量</h3>
                <div class="value">{qa.get('max_quality', 0):.3f}</div>
            </div>
            <div class="summary-card">
                <h3>高質量表格</h3>
                <div class="value">{qa.get('high_quality_count', 0)}</div>
            </div>
            <div class="summary-card">
                <h3>低質量表格</h3>
                <div class="value">{qa.get('low_quality_count', 0)}</div>
            </div>
        </div>
        """
    
    def _generate_detailed_results_section(self, all_results: Dict) -> str:
        """生成詳細結果部分"""
        html = "<h2>📋 詳細提取結果</h2>"
        
        for file_name, file_results in all_results.items():
            html += f"""
            <div class="method-section">
                <h3>📁 文件: {file_name}</h3>
            """
            
            for method, tables in file_results.items():
                if tables:
                    html += f"""
                    <h4>🔧 方法: {method} (共 {len(tables)} 個表格)</h4>
                    <table>
                        <thead>
                            <tr>
                                <th>表格編號</th>
                                <th>頁數</th>
                                <th>大小 (行×列)</th>
                                <th>質量分數</th>
                                <th>狀態</th>
                            </tr>
                        </thead>
                        <tbody>
                    """
                    
                    for table in tables:
                        quality = table['quality_score']
                        if quality > 0.7:
                            quality_class = "quality-high"
                            status = "優秀"
                        elif quality > 0.4:
                            quality_class = "quality-medium"
                            status = "良好"
                        else:
                            quality_class = "quality-low"
                            status = "需改進"
                        
                        html += f"""
                            <tr>
                                <td>{table['table_num']}</td>
                                <td>{table['page']}</td>
                                <td>{table['shape'][0]} × {table['shape'][1]}</td>
                                <td>{quality:.3f}</td>
                                <td><span class="{quality_class}">{status}</span></td>
                            </tr>
                        """
                    
                    html += """
                        </tbody>
                    </table>
                    """
            
            html += "</div>"
        
        return html
    
    def _generate_recommendations_section(self, analysis: Dict) -> str:
        """生成建議部分"""
        if 'recommendations' not in analysis:
            return ""
        
        recommendations = analysis['recommendations']
        html = """
        <h2>💡 優化建議</h2>
        <div class="recommendations">
            <ul>
        """
        
        for rec in recommendations:
            html += f"<li>🔸 {rec}</li>"
        
        html += """
            </ul>
        </div>
        """
        
        return html
    
    def _export_to_pdf_report(self, all_results: Dict, analysis: Dict) -> str:
        """導出PDF報告"""
        try:
            output_path = self.output_dir / "comprehensive_report.pdf"
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            
            story = []
            styles = getSampleStyleSheet()
            
            # 標題
            title = Paragraph("綜合PDF表格提取分析報告", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # 生成時間
            time_para = Paragraph(f"生成時間: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}", styles['Normal'])
            story.append(time_para)
            story.append(Spacer(1, 12))
            
            # 摘要部分
            if 'summary' in analysis:
                summary_title = Paragraph("提取摘要", styles['Heading2'])
                story.append(summary_title)
                
                summary_data = [
                    ['項目', '數值'],
                    ['處理文件數', str(analysis['summary'].get('total_files', 0))],
                    ['使用方法數', str(analysis['summary'].get('total_methods', 0))],
                    ['提取表格總數', str(analysis['summary'].get('total_tables', 0))],
                    ['平均表格/文件', f"{analysis['summary'].get('avg_tables_per_file', 0):.1f}"]
                ]
                
                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(summary_table)
                story.append(Spacer(1, 12))
            
            # 方法比較部分
            if 'method_comparison' in analysis:
                method_title = Paragraph("方法比較分析", styles['Heading2'])
                story.append(method_title)
                
                method_data = [['方法', '文件數', '表格數', '平均質量']]
                for method, stats in analysis['method_comparison'].items():
                    method_data.append([
                        method,
                        str(stats['files']),
                        str(stats['tables']),
                        f"{stats['avg_quality']:.3f}"
                    ])
                
                method_table = Table(method_data)
                method_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(method_table)
            
            doc.build(story)
            self.logger.info(f"PDF報告已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"PDF報告導出失敗: {e}")
            return ""
    
    def _export_to_txt(self, all_results: Dict, analysis: Dict) -> str:
        """導出TXT報告"""
        try:
            output_path = self.output_dir / "comprehensive_report.txt"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("綜合PDF表格提取分析報告\n")
                f.write("=" * 80 + "\n")
                f.write(f"生成時間: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}\n\n")
                
                # 摘要部分
                if 'summary' in analysis:
                    f.write("📊 提取摘要\n")
                    f.write("-" * 40 + "\n")
                    summary = analysis['summary']
                    f.write(f"處理文件數: {summary.get('total_files', 0)}\n")
                    f.write(f"使用方法數: {summary.get('total_methods', 0)}\n")
                    f.write(f"提取表格總數: {summary.get('total_tables', 0)}\n")
                    f.write(f"平均表格/文件: {summary.get('avg_tables_per_file', 0):.1f}\n\n")
                
                # 方法比較
                if 'method_comparison' in analysis:
                    f.write("🔍 方法比較分析\n")
                    f.write("-" * 40 + "\n")
                    for method, stats in analysis['method_comparison'].items():
                        f.write(f"{method}:\n")
                        f.write(f"  - 處理文件數: {stats['files']}\n")
                        f.write(f"  - 提取表格數: {stats['tables']}\n")
                        f.write(f"  - 平均質量: {stats['avg_quality']:.3f}\n\n")
                
                # 質量分析
                if 'quality_analysis' in analysis:
                    f.write("⭐ 質量分析\n")
                    f.write("-" * 40 + "\n")
                    qa = analysis['quality_analysis']
                    f.write(f"平均質量分數: {qa.get('mean_quality', 0):.3f}\n")
                    f.write(f"質量分數標準差: {qa.get('std_quality', 0):.3f}\n")
                    f.write(f"最高質量分數: {qa.get('max_quality', 0):.3f}\n")
                    f.write(f"最低質量分數: {qa.get('min_quality', 0):.3f}\n")
                    f.write(f"高質量表格數: {qa.get('high_quality_count', 0)}\n")
                    f.write(f"中等質量表格數: {qa.get('medium_quality_count', 0)}\n")
                    f.write(f"低質量表格數: {qa.get('low_quality_count', 0)}\n\n")
                
                # 詳細結果
                f.write("📋 詳細提取結果\n")
                f.write("=" * 60 + "\n")
                for file_name, file_results in all_results.items():
                    f.write(f"\n文件: {file_name}\n")
                    f.write("-" * 40 + "\n")
                    
                    for method, tables in file_results.items():
                        if tables:
                            f.write(f"方法 {method}: {len(tables)} 個表格\n")
                            for table in tables:
                                f.write(f"  表格 {table['table_num']}: 頁數 {table['page']}, ")
                                f.write(f"大小 {table['shape'][0]}×{table['shape'][1]}, ")
                                f.write(f"質量 {table['quality_score']:.3f}\n")
                            f.write("\n")
                
                # 建議
                if 'recommendations' in analysis:
                    f.write("💡 優化建議\n")
                    f.write("-" * 40 + "\n")
                    for i, rec in enumerate(analysis['recommendations'], 1):
                        f.write(f"{i}. {rec}\n")
            
            self.logger.info(f"TXT報告已保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"TXT報告導出失敗: {e}")
            return ""
    
    def _export_charts(self, analysis: Dict) -> List[str]:
        """導出可視化圖表"""
        try:
            chart_paths = []
            chart_dir = self.output_dir / "charts"
            chart_dir.mkdir(exist_ok=True)
            
            # 方法比較圖
            if 'method_comparison' in analysis:
                chart_path = self._create_method_comparison_chart(analysis['method_comparison'], chart_dir)
                if chart_path:
                    chart_paths.append(chart_path)
            
            # 質量分佈圖
            if 'quality_analysis' in analysis:
                chart_path = self._create_quality_distribution_chart(analysis['quality_analysis'], chart_dir)
                if chart_path:
                    chart_paths.append(chart_path)
            
            return chart_paths
            
        except Exception as e:
            self.logger.error(f"圖表導出失敗: {e}")
            return []
    
    def _create_method_comparison_chart(self, method_data: Dict, chart_dir: Path) -> str:
        """創建方法比較圖表"""
        try:
            plt.figure(figsize=(12, 6))
            
            methods = list(method_data.keys())
            qualities = [method_data[method]['avg_quality'] for method in methods]
            table_counts = [method_data[method]['tables'] for method in methods]
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
            
            # 質量比較
            bars1 = ax1.bar(methods, qualities, color='skyblue', alpha=0.8)
            ax1.set_title('各方法平均質量分數', fontsize=14, fontweight='bold')
            ax1.set_ylabel('質量分數')
            ax1.set_ylim(0, 1)
            
            # 添加數值標籤
            for bar, quality in zip(bars1, qualities):
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height + 0.01,
                        f'{quality:.3f}', ha='center', va='bottom')
            
            # 表格數量比較
            bars2 = ax2.bar(methods, table_counts, color='lightcoral', alpha=0.8)
            ax2.set_title('各方法提取表格數量', fontsize=14, fontweight='bold')
            ax2.set_ylabel('表格數量')
            
            # 添加數值標籤
            for bar, count in zip(bars2, table_counts):
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                        f'{count}', ha='center', va='bottom')
            
            # 旋轉x軸標籤
            for ax in [ax1, ax2]:
                ax.tick_params(axis='x', rotation=45)
            
            plt.tight_layout()
            
            chart_path = chart_dir / "method_comparison.png"
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(chart_path)
            
        except Exception as e:
            self.logger.error(f"方法比較圖表創建失敗: {e}")
            return ""
    
    def _create_quality_distribution_chart(self, quality_data: Dict, chart_dir: Path) -> str:
        """創建質量分佈圖表"""
        try:
            plt.figure(figsize=(10, 6))
            
            categories = ['高質量\n(>0.7)', '中等質量\n(0.4-0.7)', '低質量\n(<0.4)']
            counts = [
                quality_data.get('high_quality_count', 0),
                quality_data.get('medium_quality_count', 0),
                quality_data.get('low_quality_count', 0)
            ]
            colors = ['#27ae60', '#f39c12', '#e74c3c']
            
            # 餅圖
            plt.pie(counts, labels=categories, colors=colors, autopct='%1.1f%%', startangle=90)
            plt.title('表格質量分佈', fontsize=16, fontweight='bold')
            
            chart_path = chart_dir / "quality_distribution.png"
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(chart_path)
            
        except Exception as e:
            self.logger.error(f"質量分佈圖表創建失敗: {e}")
            return ""
    
    def _create_summary_sheet_openpyxl(self, ws, analysis: Dict):
        """創建openpyxl摘要表"""
        # 標題
        ws['A1'] = "綜合PDF表格提取分析摘要"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 基本統計
        row = 3
        if 'summary' in analysis:
            summary = analysis['summary']
            stats = [
                ("處理文件數", summary.get('total_files', 0)),
                ("使用方法數", summary.get('total_methods', 0)),
                ("提取表格總數", summary.get('total_tables', 0)),
                ("平均表格/文件", f"{summary.get('avg_tables_per_file', 0):.1f}")
            ]
            
            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
    
    def _fill_method_sheet_openpyxl(self, ws, tables: List[Dict], file_name: str, method: str):
        """填充openpyxl方法表"""
        # 標題
        ws['A1'] = f"文件: {file_name} | 方法: {method}"
        ws['A1'].font = Font(size=14, bold=True)
        
        current_row = 3
        for i, table in enumerate(tables, 1):
            # 表格標題
            ws[f'A{current_row}'] = f"表格 {i} (頁數: {table['page']}, 質量: {table['quality_score']:.3f})"
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 2
            
            # 表格數據
            df = table['data']
            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=current_row + r_idx, column=c_idx, value=value)
            
            current_row += len(df) + 3
    
    def _create_summary_sheet_xlsxwriter(self, ws, analysis: Dict, title_format, header_format):
        """創建xlsxwriter摘要表"""
        ws.write('A1', "綜合PDF表格提取分析摘要", title_format)
        
        row = 2
        if 'summary' in analysis:
            summary = analysis['summary']
            stats = [
                ("處理文件數", summary.get('total_files', 0)),
                ("使用方法數", summary.get('total_methods', 0)),
                ("提取表格總數", summary.get('total_tables', 0)),
                ("平均表格/文件", summary.get('avg_tables_per_file', 0))
            ]
            
            for label, value in stats:
                ws.write(row, 0, label, header_format)
                ws.write(row, 1, value)
                row += 1
    
    def _fill_method_sheet_xlsxwriter(self, ws, tables: List[Dict], file_name: str, method: str, header_format):
        """填充xlsxwriter方法表"""
        ws.write('A1', f"文件: {file_name} | 方法: {method}", header_format)
        
        current_row = 2
        for i, table in enumerate(tables, 1):
            ws.write(current_row, 0, f"表格 {i} (頁數: {table['page']}, 質量: {table['quality_score']:.3f})", header_format)
            current_row += 2
            
            # 寫入表格數據
            df = table['data']
            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row_data):
                    ws.write(current_row + r_idx, c_idx, value)
            
            current_row += len(df) + 3
    
    def _clean_sheet_name(self, name: str) -> str:
        """清理工作表名稱"""
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')
        return name[:31]
    
    def _clean_filename(self, name: str) -> str:
        """清理文件名"""
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for char in invalid_chars:
            name = name.replace(char, '_')
        return name


class EnhancedUserInterface:
    """增強版用戶界面"""
    
    def __init__(self, logger):
        self.logger = logger
        self.default_input_dir = DEFAULT_INPUT_DIR
        self.default_output_dir = DEFAULT_OUTPUT_DIR
    
    def run_interactive_session(self):
        """運行互動會話"""
        self.logger.info("🚀 啟動增強版互動系統")
        
        # 顯示歡迎信息
        self._display_welcome()
        
        # 檢查和安裝依賴
        lib_manager = LibraryManager()
        missing_libs = lib_manager.check_and_install_missing()
        
        if missing_libs:
            print("⚠️ 有缺失的依賴包，建議先安裝完整依賴")
            continue_anyway = input("是否繼續運行? (y/n): ").strip().lower()
            if continue_anyway != 'y':
                return
        
        # 獲取用戶配置
        config = self._get_user_configuration()
        
        # 獲取PDF文件
        pdf_files = self._get_pdf_files(config)
        
        if not pdf_files:
            print("❌ 沒有找到PDF文件")
            return
        
        # 開始處理
        self._process_files(pdf_files, config)
    
    def _display_welcome(self):
        """顯示歡迎信息"""
        print("\n" + "="*80)
        print("🚀 綜合PDF表格提取系統 v2.0")
        print("="*80)
        print("功能特色:")
        print("✅ 8種提取工具聯合作業 (pdfplumber, camelot, tabula, OCR等)")
        print("✅ JavaScript工具支持 (pdf.js, pdf2table)")  
        print("✅ 智能表格修復和質量評估")
        print("✅ 8種輸出格式 (Excel, CSV, JSON, XML, HTML, PDF, TXT, 圖表)")
        print("✅ 深度分析和優化建議")
        print("✅ 批量處理和多線程加速")
        print("="*80)
        print(f"預設輸入目錄: {self.default_input_dir}")
        print(f"預設輸出目錄: {self.default_output_dir}")
        print("="*80)
    
    def _get_user_configuration(self) -> Dict[str, Any]:
        """獲取用戶配置"""
        config = {}
        
        print("\n🔧 系統配置")
        print("-" * 40)
        
        # 輸入目錄配置
        print(f"\n📁 輸入目錄配置")
        use_default_input = self._ask_yes_no(f"使用預設輸入目錄? ({self.default_input_dir})", True)
        
        if use_default_input:
            config['input_dir'] = self.default_input_dir
        else:
            config['input_dir'] = input("請輸入PDF文件目錄: ").strip().strip('"')
        
        # 輸出目錄配置
        print(f"\n📤 輸出目錄配置")
        use_default_output = self._ask_yes_no(f"使用預設輸出目錄? ({self.default_output_dir})", True)
        
        if use_default_output:
            config['output_dir'] = self.default_output_dir
        else:
            config['output_dir'] = input("請輸入輸出目錄: ").strip().strip('"')
        
        # 提取方法選擇
        print(f"\n🔍 提取方法配置")
        print("可用方法:")
        methods = [
            "pdfplumber", "camelot", "tabula", "pdfminer", 
            "pymupdf", "ocr", "pdf.js", "pdf2table", "pdfquery"
        ]
        
        for i, method in enumerate(methods, 1):
            print(f"  {i}. {method}")
        
        use_all_methods = self._ask_yes_no("使用所有可用方法?", True)
        
        if use_all_methods:
            config['methods'] = methods
        else:
            selected = input("請選擇方法編號 (用空格分隔, 如: 1 2 3): ").strip().split()
            config['methods'] = [methods[int(i)-1] for i in selected if i.isdigit() and 1 <= int(i) <= len(methods)]
        
        # 輸出格式選擇
        print(f"\n📄 輸出格式配置")
        formats = ["excel", "csv", "json", "xml", "html", "pdf", "txt", "charts"]
        
        for i, fmt in enumerate(formats, 1):
            print(f"  {i}. {fmt}")
        
        use_all_formats = self._ask_yes_no("導出所有格式?", True)
        
        if use_all_formats:
            config['output_formats'] = formats
        else:
            selected = input("請選擇格式編號 (用空格分隔): ").strip().split()
            config['output_formats'] = [formats[int(i)-1] for i in selected if i.isdigit() and 1 <= int(i) <= len(formats)]
        
        # 其他選項
        print(f"\n⚙️ 高級選項")
        config['use_multithread'] = self._ask_yes_no("使用多線程加速?", True)
        config['fix_tables'] = self._ask_yes_no("自動修復表格?", True)
        config['deep_analysis'] = self._ask_yes_no("進行深度分析?", True)
        config['create_charts'] = self._ask_yes_no("生成可視化圖表?", True)
        
        return config
    
    def _get_pdf_files(self, config: Dict[str, Any]) -> List[str]:
        """獲取PDF文件列表"""
        input_dir = Path(config['input_dir'])
        
        if not input_dir.exists():
            print(f"❌ 輸入目錄不存在: {input_dir}")
            return []
        
        # 查找PDF文件
        pdf_files = list(input_dir.glob("*.pdf"))
        
        if not pdf_files:
            print(f"❌ 在 {input_dir} 中未找到PDF文件")
            return []
        
        print(f"\n📋 發現 {len(pdf_files)} 個PDF文件:")
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"  {i}. {pdf_file.name}")
        
        process_all = self._ask_yes_no(f"處理所有 {len(pdf_files)} 個文件?", True)
        
        if process_all:
            return [str(f) for f in pdf_files]
        else:
            selected = input("請選擇要處理的文件編號 (用空格分隔): ").strip().split()
            selected_files = []
            for i in selected:
                if i.isdigit() and 1 <= int(i) <= len(pdf_files):
                    selected_files.append(str(pdf_files[int(i)-1]))
            return selected_files
    
    def _process_files(self, pdf_files: List[str], config: Dict[str, Any]):
        """處理PDF文件"""
        print(f"\n🚀 開始處理 {len(pdf_files)} 個PDF文件...")
        
        # 創建輸出目錄
        output_dir = Path(config['output_dir'])
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 初始化組件
        logger = logging.getLogger('PDFExtractor')
        extractor = EnhancedTableExtractor(logger)
        analyzer = AdvancedAnalyzer(logger)
        exporter = ComprehensiveExporter(logger, output_dir)
        
        # 進度追蹤
        total_steps = len(pdf_files) + 2  # 文件處理 + 分析 + 導出
        current_step = 0
        
        def update_progress(description: str):
            nonlocal current_step
            current_step += 1
            percentage = (current_step / total_steps) * 100
            print(f"\r進度 [{current_step}/{total_steps}] {percentage:.1f}% - {description}", end="", flush=True)
        
        # 提取所有文件的表格
        all_results = {}
        
        if config.get('use_multithread', False) and len(pdf_files) > 1:
            print("🔄 使用多線程處理...")
            with ThreadPoolExecutor(max_workers=min(4, len(pdf_files))) as executor:
                future_to_file = {
                    executor.submit(extractor.extract_all_methods, pdf_file): pdf_file 
                    for pdf_file in pdf_files
                }
                
                for future in as_completed(future_to_file):
                    pdf_file = future_to_file[future]
                    try:
                        file_results = future.result()
                        all_results[Path(pdf_file).name] = file_results
                        update_progress(f"完成 {Path(pdf_file).name}")
                    except Exception as e:
                        logger.error(f"處理 {pdf_file} 失敗: {e}")
                        update_progress(f"失敗 {Path(pdf_file).name}")
        else:
            print("🔄 順序處理...")
            for pdf_file in pdf_files:
                try:
                    file_results = extractor.extract_all_methods(pdf_file)
                    all_results[Path(pdf_file).name] = file_results
                    update_progress(f"完成 {Path(pdf_file).name}")
                except Exception as e:
                    logger.error(f"處理 {pdf_file} 失敗: {e}")
                    update_progress(f"失敗 {Path(pdf_file).name}")
        
        print()  # 換行
        
        # 深度分析
        if config.get('deep_analysis', False):
            print("🔍 進行深度分析...")
            analysis = analyzer.analyze_all_results(all_results)
            update_progress("分析完成")
        else:
            analysis = {'summary': {'total_files': len(all_results)}}
        
        # 導出結果
        print("📤 導出結果...")
        export_paths = exporter.export_all_formats(all_results, analysis)
        update_progress("導出完成")
        
        print()  # 換行
        
        # 顯示結果摘要
        self._display_results_summary(all_results, analysis, export_paths, output_dir)
    
    def _display_results_summary(self, all_results: Dict, analysis: Dict, 
                                export_paths: Dict, output_dir: Path):
        """顯示結果摘要"""
        print("\n" + "="*80)
        print("🎉 處理完成！")
        print("="*80)
        
        # 基本統計
        total_files = len(all_results)
        total_methods = len(set().union(*(file_results.keys() for file_results in all_results.values())))
        total_tables = sum(
            len(tables) for file_results in all_results.values() 
            for tables in file_results.values()
        )
        
        print(f"📊 處理統計:")
        print(f"  • 處理文件: {total_files} 個")
        print(f"  • 使用方法: {total_methods} 種")
        print(f"  • 提取表格: {total_tables} 個")
        
        # 方法效果統計
        if 'method_comparison' in analysis:
            print(f"\n🏆 最佳方法:")
            method_comparison = analysis['method_comparison']
            best_method = max(method_comparison.keys(), 
                            key=lambda m: method_comparison[m]['avg_quality'])
            best_quality = method_comparison[best_method]['avg_quality']
            print(f"  • {best_method} (平均質量: {best_quality:.3f})")
        
        # 質量統計
        if 'quality_analysis' in analysis:
            qa = analysis['quality_analysis']
            print(f"\n⭐ 質量統計:")
            print(f"  • 平均質量: {qa.get('mean_quality', 0):.3f}")
            print(f"  • 高質量表格: {qa.get('high_quality_count', 0)} 個")
            print(f"  • 低質量表格: {qa.get('low_quality_count', 0)} 個")
        
        # 導出文件
        print(f"\n📁 導出文件:")
        for format_type, path in export_paths.items():
            if isinstance(path, list):
                print(f"  • {format_type}: {len(path)} 個文件")
            else:
                print(f"  • {format_type}: {Path(path).name}")
        
        print(f"\n📂 所有文件保存在: {output_dir}")
        
        # 建議
        if 'recommendations' in analysis:
            print(f"\n💡 優化建議:")
            for rec in analysis['recommendations']:
                print(f"  • {rec}")
        
        print("="*80)
        
        # 詢問是否打開輸出目錄
        open_dir = self._ask_yes_no("是否打開輸出目錄?", True)
        if open_dir:
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(output_dir)
                elif os.name == 'posix':  # macOS/Linux
                    os.system(f'open "{output_dir}"')
            except Exception:
                print(f"請手動打開: {output_dir}")
    
    def _ask_yes_no(self, question: str, default: bool = True) -> bool:
        """詢問是/否問題"""
        default_str = "Y/n" if default else "y/N"
        answer = input(f"{question} ({default_str}): ").strip().lower()
        
        if not answer:
            return default
        
        return answer in ['y', 'yes', '是']


def main():
    """主函數"""
    try:
        # 創建日誌器
        logger = logging.getLogger('PDFExtractor')
        logger.setLevel(logging.INFO)
        
        # 控制台處理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(levelname)s - %(message)s')
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        
        # 創建用戶界面
        ui = EnhancedUserInterface(logger)
        
        # 運行互動會話
        ui.run_interactive_session()
        
    except KeyboardInterrupt:
        print("\n\n⚠️ 用戶中斷程序")
    except Exception as e:
        print(f"\n❌ 系統錯誤: {str(e)}")
        print("請檢查依賴包是否完整安裝")


if __name__ == "__main__":
    main()


# ========================================================================
# 📋 完整功能檢查清單與安裝指南
# ========================================================================
"""
🎯 系統特色總覽
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ 8種核心提取工具: pdfplumber, camelot, tabula, pdfminer, PyMuPDF, OCR, pdfquery
✅ JavaScript工具支持: pdf.js, pdf2table (通過js2py和selenium)
✅ 智能表格修復: 不規則表格處理、合併單元格、數據清理
✅ 8種輸出格式: Excel(2種), CSV, JSON, XML, HTML, PDF, TXT, 圖表
✅ 深度質量分析: 質量評分、方法比較、統計分析
✅ 增強OCR處理: 圖像預處理、多語言識別、表格檢測
✅ 批量並行處理: 多線程加速、進度追蹤
✅ 用戶友好界面: 互動式配置、智能提示
✅ 預設路徑配置: 自動檢測VeritasReportNova目錄
✅ 依賴自動檢查: 缺失庫提示和安裝建議

📦 完整依賴安裝命令
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 基礎數據處理
pip install pandas numpy

# PDF處理工具 (核心)
pip install pdfplumber camelot-py[cv] tabula-py pdfminer.six PyMuPDF pdf2image pdfquery

# OCR和圖像處理
pip install pytesseract pillow opencv-python

# Excel和文檔輸出
pip install openpyxl xlsxwriter reportlab

# Web和JavaScript支持
pip install requests beautifulsoup4 selenium js2py

# 數據分析和可視化
pip install matplotlib seaborn scikit-learn

# 系統監控
pip install psutil

# 一鍵安裝所有依賴
pip install pandas numpy pdfplumber camelot-py[cv] tabula-py pdfminer.six PyMuPDF pdf2image pdfquery pytesseract pillow opencv-python openpyxl xlsxwriter reportlab requests beautifulsoup4 selenium js2py matplotlib seaborn scikit-learn psutil

🚀 系統啟動方式
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
python comprehensive_pdf_extractor.py

🎯 預設目錄配置
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
輸入目錄: C:/Users/tonyk/OneDrive/Desktop/VeritasReportNova/input
輸出目錄: C:/Users/tonyk/OneDrive/Desktop/VeritasReportNova/output

🏆 技術優勢
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🔧 多工具聚合: 8+種提取引擎聯合作業，互相補強
🧠 智能質量評估: 自動評分系統，識別最佳結果
🔄 自動表格修復: 處理不規則、合併單元格等問題
📊 全方位分析: 統計分析、可視化、優化建議
🌐 JavaScript整合: 支援現代Web工具，pdf.js等
⚡ 高性能處理: 多線程並行，大批量文件支持
🎨 豐富輸出格式: 8種專業格式，滿足各種需求
💻 友好用戶界面: 零學習成本，問答式配置

這是目前最完整、最強大的開源PDF表格提取解決方案！🌟
"""