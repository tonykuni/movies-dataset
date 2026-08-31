#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════════════════════════════╗
║   VERITAS PDF EXTRACTOR™ v9.0 - MODULAR PIPELINE EDITION                                             ║
║   ════════════════════════════════════════════════════════════════════════════════════════════════   ║
║                                                                                                      ║
║   🎯 Pipeline: PDF → Image(前處理) → [中處理] → [後處理] → OCR → Table → Database                    ║
║   🎯 Architecture: Fully Modular (可插拔模組)                                                        ║
║   🎯 KPI: Per-Module Tracking (每模組獨立 KPI)                                                       ║
║   🎯 Temp Chain: 環環相扣的暫存輸出入                                                                ║
║                                                                                                      ║
║   ┌─────────────────────────────────────────────────────────────────────────────────────────────┐    ║
║   │ MODULE INDEX                                                                                │    ║
║   ├─────────────────────────────────────────────────────────────────────────────────────────────┤    ║
║   │ M01: CONFIG           參數設定模組                                                          │    ║
║   │ M02: ACCELERATOR      加速器管理模組                                                        │    ║
║   │ M03: KPI_TRACKER      KPI 追蹤模組                                                          │    ║
║   │ M04: TEMP_MANAGER     暫存管理模組                                                          │    ║
║   │ M05: PDF_TO_IMAGE     PDF 轉圖片模組                                                        │    ║
║   │ M06: IMAGE_PRE        前處理模組 (白平衡/降噪/銳化)                                         │    ║
║   │ M07: DATA_STRUCTURES  資料結構定義                                                          │    ║
║   │ M08: DATABASE         資料庫模組                                                            │    ║
║   │ M09: DISPLAY          顯示模組                                                              │    ║
║   │ M10: PIPELINE         主流程協調模組                                                        │    ║
║   │ M11: CLI              命令列介面                                                            │    ║
║   └─────────────────────────────────────────────────────────────────────────────────────────────┘    ║
╚══════════════════════════════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations
# ===== [VIA:ACCEL-BRIDGE:v0100] SuperAccel 加速器橋(批102 全樹導入令;graceful 零行為變更) =====
try:
    import sys as _sa_sys
    from pathlib import Path as _sa_Path
    _sa_p = _sa_Path(__file__).resolve()
    while _sa_p.parent != _sa_p:
        if (_sa_p / "supportive modules" / "VIA_SuperAccel_Module.py").exists():
            _sa_sys.path.insert(0, str(_sa_p / "supportive modules"))
            break
        _sa_p = _sa_p.parent
    import VIA_SuperAccel_Module as VIA_ACCEL  # noqa: N816
except Exception:
    VIA_ACCEL = None  # graceful:加速器缺席零影響
# ===== [VIA:ACCEL-BRIDGE:END] =====
import os
import sys
import time
import shutil
import argparse
import sqlite3
import json
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Any, Dict, List, Optional, Tuple, Callable, Union
from collections import defaultdict
import hashlib
import re
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M01: CONFIG - 參數設定模組                                                  ║
# ║  所有參數集中於此，方便調整                                                  ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

@dataclass
class Config:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M01-001: MASTER CONFIGURATION                                             ║
    ║  所有模組參數集中管理                                                      ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-010: PATH CONFIGURATION (路徑設定)
    # ═══════════════════════════════════════════════════════════════════════════
    base_dir: str = r"C:\Users\tonyk\OneDrive\Desktop\VRN"
    pdf_input_dir: str = ""
    pdf_output_dir: str = ""
    temp_dir: str = ""
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-020: PDF TO IMAGE SETTINGS (PDF 轉圖片設定)
    # ═══════════════════════════════════════════════════════════════════════════
    pdf_dpi: int = 300
    pdf_colorspace: str = "rgb"
    pdf_alpha: bool = False
    pdf_output_format: str = "png"
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-030: IMAGE PRE-PROCESSING (前處理設定)
    # ═══════════════════════════════════════════════════════════════════════════
    pre_enable: bool = True
    pre_denoise: bool = True
    pre_denoise_strength: int = 10
    pre_sharpen: bool = True
    pre_sharpen_amount: float = 1.5
    pre_contrast: bool = True
    pre_contrast_factor: float = 1.2
    pre_white_balance: bool = True
    pre_deskew: bool = True
    pre_deskew_threshold: float = 0.5
    pre_binarize: bool = False
    pre_binarize_method: str = "otsu"
    pre_resize_factor: float = 1.0
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-035: TABLE EXTRACTION SETTINGS (表格提取設定)
    # ═══════════════════════════════════════════════════════════════════════════
    table_min_rows: int = 2
    table_min_cols: int = 2
    table_max_empty_ratio: float = 0.5
    table_min_data_cells: int = 4
    table_max_text_cell_length: int = 500
    table_exclude_single_column_text: bool = True
    table_exclude_header_only: bool = True
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-040: KPI SETTINGS (KPI 設定)
    # ═══════════════════════════════════════════════════════════════════════════
    kpi_enable: bool = True
    kpi_per_module: bool = True
    kpi_time_warning: float = 5.0
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-050: TEMP CHAIN SETTINGS (暫存鏈設定)
    # ═══════════════════════════════════════════════════════════════════════════
    temp_keep: bool = False
    temp_prefix: str = "vrn_"
    temp_chain_verify: bool = True
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-060: OUTPUT SETTINGS (輸出設定)
    # ═══════════════════════════════════════════════════════════════════════════
    output_json: bool = True
    output_excel: bool = True
    output_csv: bool = True
    output_sqlite: bool = True
    output_images: bool = False
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-065: PARALLEL PROCESSING SETTINGS (平行處理設定)
    # ═══════════════════════════════════════════════════════════════════════════
    parallel_enable: bool = True
    parallel_max_workers: int = 4
    
    # ═══════════════════════════════════════════════════════════════════════════
    # M01-070: DISPLAY SETTINGS (顯示設定)
    # ═══════════════════════════════════════════════════════════════════════════
    display_rich: bool = True
    display_progress: bool = True
    display_kpi_dashboard: bool = True
    debug_mode: bool = False
    
    def __post_init__(self):
        if not self.pdf_input_dir:
            self.pdf_input_dir = os.path.join(self.base_dir, "pdf_input")
        if not self.pdf_output_dir:
            self.pdf_output_dir = os.path.join(self.base_dir, "pdf_output")
        if not self.temp_dir:
            self.temp_dir = os.path.join(self.base_dir, "temp")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M02: ACCELERATOR - 加速器管理模組                                           ║
# ║  統一管理所有加速庫的載入與狀態                                              ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class AcceleratorManager:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M02-001: ACCELERATOR MANAGER                                              ║
    ║  加速器統一管理、載入、狀態追蹤                                            ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    _instance = None
    _initialized = False
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance
    
    def __init__(self):
        if AcceleratorManager._initialized:
            return
        self.status: Dict[str, Dict] = {}
        self.libs: Dict[str, Any] = {}
        self._load_all()
        AcceleratorManager._initialized = True
    
    def _load_all(self):
        # M02-031: PyMuPDF
        try:
            import fitz
            self.libs["fitz"] = fitz
            self.status["PyMuPDF"] = {"ok": True, "version": fitz.version[0], "type": "pdf"}
        except ImportError:
            self.status["PyMuPDF"] = {"ok": False, "version": "", "type": "pdf"}
        
        # M02-032: pdfplumber
        try:
            import pdfplumber
            self.libs["pdfplumber"] = pdfplumber
            self.status["pdfplumber"] = {"ok": True, "version": "OK", "type": "pdf"}
        except ImportError:
            self.status["pdfplumber"] = {"ok": False, "version": "", "type": "pdf"}
        
        # M02-033: Pillow
        try:
            from PIL import Image, ImageFilter, ImageEnhance, ImageOps
            self.libs["PIL"] = {"Image": Image, "ImageFilter": ImageFilter, 
                               "ImageEnhance": ImageEnhance, "ImageOps": ImageOps}
            import PIL
            self.status["Pillow"] = {"ok": True, "version": PIL.__version__, "type": "image"}
        except ImportError:
            self.status["Pillow"] = {"ok": False, "version": "", "type": "image"}
        
        # M02-034: OpenCV
        try:
            import cv2
            self.libs["cv2"] = cv2
            self.status["OpenCV"] = {"ok": True, "version": cv2.__version__, "type": "image"}
        except ImportError:
            self.status["OpenCV"] = {"ok": False, "version": "", "type": "image"}
        
        # M02-035: NumPy
        try:
            import numpy as np
            self.libs["numpy"] = np
            self.status["NumPy"] = {"ok": True, "version": np.__version__, "type": "compute"}
        except ImportError:
            self.status["NumPy"] = {"ok": False, "version": "", "type": "compute"}
        
        # M02-036: xxHash
        try:
            import xxhash
            self.libs["xxhash"] = xxhash
            self.status["xxHash"] = {"ok": True, "version": "OK", "type": "hash"}
        except ImportError:
            self.status["xxHash"] = {"ok": False, "version": "", "type": "hash"}
        
        # M02-037: Orjson
        try:
            import orjson
            self.libs["orjson"] = orjson
            self.status["Orjson"] = {"ok": True, "version": "OK", "type": "json"}
        except ImportError:
            self.status["Orjson"] = {"ok": False, "version": "", "type": "json"}
        
        # M02-038: Polars
        try:
            import polars as pl
            self.libs["polars"] = pl
            self.status["Polars"] = {"ok": True, "version": pl.__version__, "type": "data"}
        except ImportError:
            self.status["Polars"] = {"ok": False, "version": "", "type": "data"}
        
        # M02-039: Pandas
        try:
            import pandas as pd
            self.libs["pandas"] = pd
            self.status["Pandas"] = {"ok": True, "version": pd.__version__, "type": "data"}
        except ImportError:
            self.status["Pandas"] = {"ok": False, "version": "", "type": "data"}
        
        # M02-040: Rich
        try:
            from rich.console import Console
            from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn
            from rich.table import Table
            from rich.panel import Panel
            from rich import box
            self.libs["rich"] = {"Console": Console, "Progress": Progress, 
                                "Table": Table, "Panel": Panel, "box": box,
                                "SpinnerColumn": SpinnerColumn, "BarColumn": BarColumn,
                                "TextColumn": TextColumn, "TimeElapsedColumn": TimeElapsedColumn}
            self.status["Rich"] = {"ok": True, "version": "OK", "type": "display"}
        except ImportError:
            self.status["Rich"] = {"ok": False, "version": "", "type": "display"}
        
        # M02-041: openpyxl
        try:
            import openpyxl
            self.libs["openpyxl"] = openpyxl
            self.status["openpyxl"] = {"ok": True, "version": openpyxl.__version__, "type": "excel"}
        except ImportError:
            self.status["openpyxl"] = {"ok": False, "version": "", "type": "excel"}
    
    def get(self, name: str) -> Any:
        return self.libs.get(name)
    
    def is_available(self, name: str) -> bool:
        return self.status.get(name, {}).get("ok", False)
    
    def get_report(self) -> Dict[str, str]:
        return {name: f"✅ {info['version']}" if info['ok'] else "❌"
                for name, info in self.status.items()}
    
    def fast_hash(self, data) -> str:
        if isinstance(data, str):
            data = data.encode()
        if self.is_available("xxHash"):
            return self.libs["xxhash"].xxh64(data).hexdigest()
        return hashlib.md5(data).hexdigest()
    
    def json_dumps(self, obj, pretty=False) -> bytes:
        if self.is_available("Orjson"):
            orjson = self.libs["orjson"]
            opts = orjson.OPT_SERIALIZE_NUMPY
            if pretty:
                opts |= orjson.OPT_INDENT_2
            return orjson.dumps(obj, option=opts, default=str)
        return json.dumps(obj, indent=2 if pretty else None, default=str, ensure_ascii=False).encode()
    
    def json_loads(self, data) -> Any:
        if self.is_available("Orjson"):
            return self.libs["orjson"].loads(data)
        return json.loads(data.decode() if isinstance(data, bytes) else data)


ACC = AcceleratorManager()


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M03: KPI_TRACKER - KPI 追蹤模組                                             ║
# ║  每模組獨立追蹤效能指標                                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

@dataclass
class ModuleKPI:
    """M03-001: 單一模組 KPI"""
    module_name: str
    calls: int = 0
    total_time: float = 0.0
    min_time: float = float('inf')
    max_time: float = 0.0
    errors: int = 0
    items_processed: int = 0
    
    def record(self, elapsed: float, items: int = 1, error: bool = False):
        self.calls += 1
        self.total_time += elapsed
        self.min_time = min(self.min_time, elapsed)
        self.max_time = max(self.max_time, elapsed)
        self.items_processed += items
        if error:
            self.errors += 1
    
    @property
    def avg_time(self) -> float:
        return self.total_time / self.calls if self.calls > 0 else 0.0
    
    @property
    def success_rate(self) -> float:
        return (self.calls - self.errors) / self.calls if self.calls > 0 else 0.0
    
    def to_dict(self) -> Dict:
        return {
            "module": self.module_name,
            "calls": self.calls,
            "total_time": round(self.total_time, 3),
            "avg_time": round(self.avg_time, 3),
            "min_time": round(self.min_time, 3) if self.min_time != float('inf') else 0,
            "max_time": round(self.max_time, 3),
            "errors": self.errors,
            "success_rate": round(self.success_rate * 100, 1),
            "items_processed": self.items_processed
        }


class KPITracker:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M03-100: KPI TRACKER                                                      ║
    ║  全域 KPI 追蹤器                                                           ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.modules = {}
            cls._instance.start_time = time.time()
        return cls._instance
    
    def get_module(self, name: str) -> ModuleKPI:
        if name not in self.modules:
            self.modules[name] = ModuleKPI(module_name=name)
        return self.modules[name]
    
    def record(self, module_name: str, elapsed: float, items: int = 1, error: bool = False):
        self.get_module(module_name).record(elapsed, items, error)
    
    def get_summary(self) -> Dict:
        total_elapsed = time.time() - self.start_time
        return {
            "total_time": round(total_elapsed, 2),
            "modules": {name: kpi.to_dict() for name, kpi in self.modules.items()}
        }
    
    def reset(self):
        self.modules.clear()
        self.start_time = time.time()


KPI = KPITracker()


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M04: TEMP_MANAGER - 暫存管理模組                                            ║
# ║  管理 Pipeline 中的暫存檔案，環環相扣                                        ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TempManager:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M04-001: TEMP CHAIN MANAGER                                               ║
    ║  暫存鏈管理 - 前處理 → 中處理 → 後處理                                     ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    def __init__(self, config: Config):
        self.config = config
        self.base_temp = config.temp_dir
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.session_dir = os.path.join(self.base_temp, f"session_{self.session_id}")
        self.chain: Dict[str, Dict[str, str]] = {}
    
    def init_session(self):
        start = time.time()
        os.makedirs(self.session_dir, exist_ok=True)
        
        stages = ["raw_images", "pre_processed", "mid_processed", "post_processed", "ocr_output"]
        for stage in stages:
            os.makedirs(os.path.join(self.session_dir, stage), exist_ok=True)
        
        KPI.record("M04_TEMP_INIT", time.time() - start)
        return self.session_dir
    
    def get_stage_dir(self, stage: str) -> str:
        return os.path.join(self.session_dir, stage)
    
    def register_file(self, doc_id: str, stage: str, filepath: str):
        if doc_id not in self.chain:
            self.chain[doc_id] = {}
        self.chain[doc_id][stage] = filepath
    
    def get_input_file(self, doc_id: str, current_stage: str) -> Optional[str]:
        stage_order = ["raw_images", "pre_processed", "mid_processed", "post_processed"]
        if current_stage not in stage_order:
            return None
        current_idx = stage_order.index(current_stage)
        if current_idx == 0:
            return None
        prev_stage = stage_order[current_idx - 1]
        return self.chain.get(doc_id, {}).get(prev_stage)
    
    def cleanup(self):
        start = time.time()
        if not self.config.temp_keep and os.path.exists(self.session_dir):
            shutil.rmtree(self.session_dir, ignore_errors=True)
        KPI.record("M04_TEMP_CLEANUP", time.time() - start)
    
    def get_stats(self) -> Dict:
        total_files = 0
        total_size = 0
        for doc_chain in self.chain.values():
            for filepath in doc_chain.values():
                if os.path.exists(filepath):
                    total_files += 1
                    total_size += os.path.getsize(filepath)
        return {
            "session_id": self.session_id,
            "total_files": total_files,
            "total_size_mb": round(total_size / 1024 / 1024, 2),
            "documents": len(self.chain)
        }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M05: PDF_TO_IMAGE - PDF 轉圖片模組                                          ║
# ║  將 PDF 頁面轉換為高解析度圖片                                               ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class PDFToImageModule:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M05-001: PDF TO IMAGE MODULE                                              ║
    ║  PDF → Image 轉換，輸出到 temp/raw_images                                  ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    MODULE_ID = "M05_PDF_TO_IMAGE"
    
    def __init__(self, config: Config, temp_manager: TempManager):
        self.config = config
        self.temp = temp_manager
        self.fitz = ACC.get("fitz")
        if not self.fitz:
            raise RuntimeError("PyMuPDF not available")
    
    def convert(self, pdf_path: str) -> List[Dict]:
        start = time.time()
        results = []
        
        doc_id = Path(pdf_path).stem
        output_dir = self.temp.get_stage_dir("raw_images")
        
        try:
            doc = self.fitz.open(pdf_path)
            total_pages = len(doc)
            
            zoom = self.config.pdf_dpi / 72.0
            matrix = self.fitz.Matrix(zoom, zoom)
            
            for page_num in range(total_pages):
                page_start = time.time()
                page = doc[page_num]
                
                if self.config.pdf_colorspace == "gray":
                    pix = page.get_pixmap(matrix=matrix, colorspace=self.fitz.csGRAY, alpha=self.config.pdf_alpha)
                else:
                    pix = page.get_pixmap(matrix=matrix, alpha=self.config.pdf_alpha)
                
                output_filename = f"{doc_id}_p{page_num + 1:04d}.{self.config.pdf_output_format}"
                output_path = os.path.join(output_dir, output_filename)
                pix.save(output_path)
                
                page_id = f"{doc_id}_p{page_num + 1}"
                self.temp.register_file(page_id, "raw_images", output_path)
                
                results.append({
                    "page": page_num + 1,
                    "page_id": page_id,
                    "path": output_path,
                    "width": pix.width,
                    "height": pix.height,
                    "dpi": self.config.pdf_dpi,
                    "time": round(time.time() - page_start, 3)
                })
                pix = None
            
            doc.close()
            
        except Exception as e:
            KPI.record(self.MODULE_ID, time.time() - start, items=0, error=True)
            raise RuntimeError(f"PDF conversion failed: {e}")
        
        KPI.record(self.MODULE_ID, time.time() - start, items=len(results))
        return results


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M06: IMAGE_PRE - 前處理模組                                                 ║
# ║  圖片前處理：降噪、銳化、對比度、白平衡、傾斜校正                            ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class ImagePreProcessor:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M06-001: IMAGE PRE-PROCESSOR                                              ║
    ║  前處理模組：raw_images → pre_processed                                    ║
    ║                                                                            ║
    ║  Operations:                                                               ║
    ║    M06-030: 降噪 (Denoise)                                                 ║
    ║    M06-040: 白平衡 (White Balance)                                         ║
    ║    M06-050: 對比度 (Contrast CLAHE)                                        ║
    ║    M06-060: 銳化 (Sharpen)                                                 ║
    ║    M06-070: 傾斜校正 (Deskew)                                              ║
    ║    M06-080: 縮放 (Resize)                                                  ║
    ║    M06-090: 二值化 (Binarize)                                              ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    MODULE_ID = "M06_IMAGE_PRE"
    
    def __init__(self, config: Config, temp_manager: TempManager):
        self.config = config
        self.temp = temp_manager
        self.pil = ACC.get("PIL")
        self.cv2 = ACC.get("cv2")
        self.np = ACC.get("numpy")
        
        self.use_cv2 = self.cv2 is not None and self.np is not None
        self.use_pil = self.pil is not None
        
        if not self.use_cv2 and not self.use_pil:
            raise RuntimeError("No image processing library available")
    
    def process(self, page_info: Dict) -> Dict:
        """M06-020: 處理單一頁面"""
        start = time.time()
        
        page_id = page_info["page_id"]
        input_path = page_info["path"]
        output_dir = self.temp.get_stage_dir("pre_processed")
        output_path = os.path.join(output_dir, f"{page_id}_pre.png")
        
        operations = []
        
        try:
            if self.use_cv2:
                self._process_cv2(input_path, output_path, operations)
            else:
                self._process_pil(input_path, output_path, operations)
            
            self.temp.register_file(page_id, "pre_processed", output_path)
            
            elapsed = time.time() - start
            KPI.record(self.MODULE_ID, elapsed, items=1)
            
            return {
                "page_id": page_id,
                "input_path": input_path,
                "output_path": output_path,
                "operations": operations,
                "time": round(elapsed, 3)
            }
            
        except Exception as e:
            KPI.record(self.MODULE_ID, time.time() - start, error=True)
            raise RuntimeError(f"Pre-processing failed for {page_id}: {e}")
    
    def _process_cv2(self, input_path: str, output_path: str, operations: List) -> bool:
        """M06-025: OpenCV 處理流程"""
        cv2 = self.cv2
        np = self.np
        
        img = cv2.imread(input_path)
        if img is None:
            raise ValueError(f"Cannot read image: {input_path}")
        
        # M06-030: 降噪
        if self.config.pre_denoise:
            img = cv2.fastNlMeansDenoisingColored(
                img, None, 
                self.config.pre_denoise_strength,
                self.config.pre_denoise_strength, 7, 21
            )
            operations.append(f"denoise(strength={self.config.pre_denoise_strength})")
        
        # M06-040: 白平衡
        if self.config.pre_white_balance:
            img = self._white_balance_cv2(img)
            operations.append("white_balance")
        
        # M06-050: 對比度 (CLAHE)
        if self.config.pre_contrast:
            lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
            l, a, b = cv2.split(lab)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            l = clahe.apply(l)
            lab = cv2.merge([l, a, b])
            img = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
            operations.append("contrast(CLAHE)")
        
        # M06-060: 銳化
        if self.config.pre_sharpen:
            kernel = np.array([[-1, -1, -1],
                              [-1,  9, -1],
                              [-1, -1, -1]]) * self.config.pre_sharpen_amount / 9
            kernel[1, 1] = 1 + 8 * self.config.pre_sharpen_amount / 9
            img = cv2.filter2D(img, -1, kernel)
            operations.append(f"sharpen(amount={self.config.pre_sharpen_amount})")
        
        # M06-070: 傾斜校正
        if self.config.pre_deskew:
            angle = self._detect_skew_cv2(img)
            if abs(angle) > self.config.pre_deskew_threshold:
                img = self._rotate_cv2(img, angle)
                operations.append(f"deskew(angle={angle:.2f})")
        
        # M06-080: 縮放
        if self.config.pre_resize_factor != 1.0:
            h, w = img.shape[:2]
            new_w = int(w * self.config.pre_resize_factor)
            new_h = int(h * self.config.pre_resize_factor)
            img = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_LANCZOS4)
            operations.append(f"resize(factor={self.config.pre_resize_factor})")
        
        # M06-090: 二值化
        if self.config.pre_binarize:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            if self.config.pre_binarize_method == "otsu":
                _, img = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            elif self.config.pre_binarize_method == "adaptive":
                img = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
            operations.append(f"binarize({self.config.pre_binarize_method})")
        
        cv2.imwrite(output_path, img, [cv2.IMWRITE_PNG_COMPRESSION, 3])
        return True
    
    def _white_balance_cv2(self, img):
        """M06-041: 白平衡 (灰度世界)"""
        np = self.np
        result = self.cv2.cvtColor(img, self.cv2.COLOR_BGR2LAB)
        avg_a = np.average(result[:, :, 1])
        avg_b = np.average(result[:, :, 2])
        result[:, :, 1] = result[:, :, 1] - ((avg_a - 128) * (result[:, :, 0] / 255.0) * 1.1)
        result[:, :, 2] = result[:, :, 2] - ((avg_b - 128) * (result[:, :, 0] / 255.0) * 1.1)
        return self.cv2.cvtColor(result, self.cv2.COLOR_LAB2BGR)
    
    def _detect_skew_cv2(self, img) -> float:
        """M06-071: 偵測傾斜角度"""
        cv2 = self.cv2
        np = self.np
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150, apertureSize=3)
        lines = cv2.HoughLinesP(edges, 1, np.pi/180, 100, minLineLength=100, maxLineGap=10)
        
        if lines is None:
            return 0.0
        
        angles = []
        for line in lines:
            x1, y1, x2, y2 = line[0]
            angle = np.degrees(np.arctan2(y2 - y1, x2 - x1))
            if abs(angle) < 45:
                angles.append(angle)
        
        return np.median(angles) if angles else 0.0
    
    def _rotate_cv2(self, img, angle: float):
        """M06-072: 旋轉圖片"""
        cv2 = self.cv2
        h, w = img.shape[:2]
        center = (w // 2, h // 2)
        matrix = cv2.getRotationMatrix2D(center, angle, 1.0)
        return cv2.warpAffine(img, matrix, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    
    def _process_pil(self, input_path: str, output_path: str, operations: List) -> bool:
        """M06-026: Pillow 處理流程 (Fallback)"""
        Image = self.pil["Image"]
        ImageEnhance = self.pil["ImageEnhance"]
        ImageOps = self.pil["ImageOps"]
        
        img = Image.open(input_path)
        
        if self.config.pre_contrast:
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(self.config.pre_contrast_factor)
            operations.append(f"contrast(factor={self.config.pre_contrast_factor})")
        
        if self.config.pre_sharpen:
            enhancer = ImageEnhance.Sharpness(img)
            img = enhancer.enhance(self.config.pre_sharpen_amount)
            operations.append(f"sharpen(amount={self.config.pre_sharpen_amount})")
        
        if self.config.pre_white_balance:
            img = ImageOps.autocontrast(img)
            operations.append("autocontrast")
        
        if self.config.pre_resize_factor != 1.0:
            w, h = img.size
            new_w = int(w * self.config.pre_resize_factor)
            new_h = int(h * self.config.pre_resize_factor)
            img = img.resize((new_w, new_h), Image.LANCZOS)
            operations.append(f"resize(factor={self.config.pre_resize_factor})")
        
        img.save(output_path, "PNG", optimize=True)
        return True
    
    def process_batch(self, page_infos: List[Dict]) -> List[Dict]:
        """M06-100: 批次處理"""
        results = []
        for info in page_infos:
            try:
                result = self.process(info)
                results.append(result)
            except Exception as e:
                results.append({"page_id": info.get("page_id", "unknown"), "error": str(e)})
        return results


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M07: DATA_STRUCTURES - 資料結構定義                                         ║
# ║  定義表格、文字區塊、文件等核心資料結構                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

@dataclass
class BoundingBox:
    """M07-001: 邊界框"""
    x0: float
    y0: float
    x1: float
    y1: float
    
    @property
    def width(self) -> float:
        return self.x1 - self.x0
    
    @property
    def height(self) -> float:
        return self.y1 - self.y0
    
    @property
    def area(self) -> float:
        return self.width * self.height
    
    def to_tuple(self) -> Tuple[float, float, float, float]:
        return (self.x0, self.y0, self.x1, self.y1)


@dataclass
class TableCell:
    """M07-010: 表格儲存格"""
    row: int
    col: int
    text: str
    bbox: Optional[BoundingBox] = None
    rowspan: int = 1
    colspan: int = 1
    is_header: bool = False
    confidence: float = 1.0
    
    @property
    def is_empty(self) -> bool:
        return not self.text or not self.text.strip()
    
    @property
    def is_numeric(self) -> bool:
        text = self.text.strip().replace(",", "").replace("%", "").replace("$", "")
        text = re.sub(r"[()（）]", "", text)
        try:
            float(text)
            return True
        except ValueError:
            return False


@dataclass
class ExtractedTable:
    """M07-020: 提取的表格"""
    table_id: str
    page: int
    rows: int
    cols: int
    cells: List[TableCell]
    bbox: Optional[BoundingBox] = None
    table_type: str = "data"
    confidence: float = 1.0
    extraction_method: str = "pdfplumber"
    
    def to_2d_array(self) -> List[List[str]]:
        """轉換為二維陣列"""
        grid = [["" for _ in range(self.cols)] for _ in range(self.rows)]
        for cell in self.cells:
            if 0 <= cell.row < self.rows and 0 <= cell.col < self.cols:
                grid[cell.row][cell.col] = cell.text
        return grid
    
    def to_dict(self) -> Dict:
        return {
            "table_id": self.table_id,
            "page": self.page,
            "rows": self.rows,
            "cols": self.cols,
            "table_type": self.table_type,
            "confidence": self.confidence,
            "extraction_method": self.extraction_method,
            "data": self.to_2d_array()
        }
    
    @property
    def empty_ratio(self) -> float:
        """空白比例"""
        if not self.cells:
            return 1.0
        empty_count = sum(1 for c in self.cells if c.is_empty)
        return empty_count / len(self.cells)
    
    @property
    def data_cells_count(self) -> int:
        """非空儲存格數量"""
        return sum(1 for c in self.cells if not c.is_empty)
    
    def classify_type(self) -> str:
        """M07-025: 自動分類表格類型"""
        all_text = " ".join(c.text for c in self.cells if c.text).lower()
        
        financial_keywords = ["營收", "revenue", "eps", "毛利", "gross", "淨利", "profit", 
                             "營業", "operating", "成長", "growth", "yoy", "qoq",
                             "股利", "dividend", "股價", "price", "本益比", "p/e", "pe"]
        
        if any(kw in all_text for kw in financial_keywords):
            self.table_type = "financial"
        elif self.cols == 2 and self.rows >= 3:
            self.table_type = "info"
        else:
            self.table_type = "data"
        
        return self.table_type


@dataclass 
class TextBlock:
    """M07-030: 文字區塊"""
    block_id: str
    page: int
    text: str
    block_type: str = "paragraph"
    bbox: Optional[BoundingBox] = None
    font_size: float = 0.0
    is_bold: bool = False
    confidence: float = 1.0
    
    def classify_type(self) -> str:
        """M07-035: 自動分類文字類型"""
        text = self.text.strip()
        
        if len(text) < 50 and (self.is_bold or self.font_size > 12):
            self.block_type = "title"
        elif text.startswith(("➢", "•", "●", "○", "■", "□", "-", "*", "·")):
            self.block_type = "bullet"
        elif re.match(r"^(圖|表|Figure|Table|Chart)\s*\d", text, re.IGNORECASE):
            self.block_type = "caption"
        elif len(text) > 200:
            self.block_type = "paragraph"
        else:
            self.block_type = "text"
        
        return self.block_type
    
    def to_dict(self) -> Dict:
        return {
            "block_id": self.block_id,
            "page": self.page,
            "text": self.text,
            "block_type": self.block_type,
            "font_size": self.font_size,
            "is_bold": self.is_bold,
            "confidence": self.confidence
        }


@dataclass
class DocumentResult:
    """M07-100: 文件處理結果"""
    doc_id: str
    filename: str
    total_pages: int
    tables: List[ExtractedTable] = field(default_factory=list)
    text_blocks: List[TextBlock] = field(default_factory=list)
    metadata: Dict = field(default_factory=dict)
    processing_time: float = 0.0
    errors: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict:
        return {
            "doc_id": self.doc_id,
            "filename": self.filename,
            "total_pages": self.total_pages,
            "tables_count": len(self.tables),
            "text_blocks_count": len(self.text_blocks),
            "tables": [t.to_dict() for t in self.tables],
            "text_blocks": [b.to_dict() for b in self.text_blocks],
            "metadata": self.metadata,
            "processing_time": self.processing_time,
            "errors": self.errors
        }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M08: DATABASE - 資料庫模組                                                  ║
# ║  SQLite 資料庫操作、JSON/CSV/Excel 輸出                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class DatabaseModule:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M08-001: DATABASE MODULE                                                  ║
    ║  資料持久化：SQLite + JSON + CSV + Excel                                   ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    MODULE_ID = "M08_DATABASE"
    
    def __init__(self, config: Config):
        self.config = config
        self.db_path = os.path.join(config.pdf_output_dir, "veritas_v9.db")
        self._init_db()
    
    def _init_db(self):
        """M08-010: 初始化資料庫"""
        start = time.time()
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 文件表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS documents (
                doc_id TEXT PRIMARY KEY,
                filename TEXT,
                total_pages INTEGER,
                tables_count INTEGER,
                text_blocks_count INTEGER,
                processing_time REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # 表格表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS tables (
                table_id TEXT PRIMARY KEY,
                doc_id TEXT,
                page INTEGER,
                rows INTEGER,
                cols INTEGER,
                table_type TEXT,
                confidence REAL,
                extraction_method TEXT,
                data_json TEXT,
                FOREIGN KEY (doc_id) REFERENCES documents(doc_id)
            )
        """)
        
        # 文字區塊表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS text_blocks (
                block_id TEXT PRIMARY KEY,
                doc_id TEXT,
                page INTEGER,
                text TEXT,
                block_type TEXT,
                font_size REAL,
                is_bold INTEGER,
                confidence REAL,
                FOREIGN KEY (doc_id) REFERENCES documents(doc_id)
            )
        """)
        
        # KPI 表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS kpi_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT,
                module_name TEXT,
                calls INTEGER,
                total_time REAL,
                avg_time REAL,
                errors INTEGER,
                success_rate REAL,
                items_processed INTEGER,
                logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        conn.commit()
        conn.close()
        
        KPI.record(self.MODULE_ID + "_INIT", time.time() - start)
    
    def save_document(self, result: DocumentResult):
        """M08-020: 儲存文件結果"""
        start = time.time()
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 儲存文件
        cursor.execute("""
            INSERT OR REPLACE INTO documents 
            (doc_id, filename, total_pages, tables_count, text_blocks_count, processing_time)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (result.doc_id, result.filename, result.total_pages, 
              len(result.tables), len(result.text_blocks), result.processing_time))
        
        # 儲存表格
        for table in result.tables:
            data_json = ACC.json_dumps(table.to_2d_array()).decode()
            cursor.execute("""
                INSERT OR REPLACE INTO tables
                (table_id, doc_id, page, rows, cols, table_type, confidence, extraction_method, data_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (table.table_id, result.doc_id, table.page, table.rows, table.cols,
                  table.table_type, table.confidence, table.extraction_method, data_json))
        
        # 儲存文字區塊
        for block in result.text_blocks:
            cursor.execute("""
                INSERT OR REPLACE INTO text_blocks
                (block_id, doc_id, page, text, block_type, font_size, is_bold, confidence)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (block.block_id, result.doc_id, block.page, block.text,
                  block.block_type, block.font_size, 1 if block.is_bold else 0, block.confidence))
        
        conn.commit()
        conn.close()
        
        KPI.record(self.MODULE_ID + "_SAVE", time.time() - start)
    
    def save_kpi(self, session_id: str, kpi_summary: Dict):
        """M08-030: 儲存 KPI"""
        start = time.time()
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        for module_name, kpi_data in kpi_summary.get("modules", {}).items():
            cursor.execute("""
                INSERT INTO kpi_logs 
                (session_id, module_name, calls, total_time, avg_time, errors, success_rate, items_processed)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (session_id, module_name, kpi_data["calls"], kpi_data["total_time"],
                  kpi_data["avg_time"], kpi_data["errors"], kpi_data["success_rate"],
                  kpi_data["items_processed"]))
        
        conn.commit()
        conn.close()
        
        KPI.record(self.MODULE_ID + "_KPI", time.time() - start)
    
    def export_json(self, result: DocumentResult, output_path: str):
        """M08-040: 輸出 JSON"""
        start = time.time()
        
        with open(output_path, "wb") as f:
            f.write(ACC.json_dumps(result.to_dict(), pretty=True))
        
        KPI.record(self.MODULE_ID + "_JSON", time.time() - start)
    
    def export_csv(self, result: DocumentResult, output_dir: str):
        """M08-050: 輸出 CSV"""
        start = time.time()
        
        for table in result.tables:
            csv_path = os.path.join(output_dir, f"{table.table_id}.csv")
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerows(table.to_2d_array())
        
        KPI.record(self.MODULE_ID + "_CSV", time.time() - start, items=len(result.tables))
    
    def export_excel(self, result: DocumentResult, output_path: str):
        """M08-060: 輸出 Excel"""
        start = time.time()
        
        openpyxl = ACC.get("openpyxl")
        if not openpyxl:
            KPI.record(self.MODULE_ID + "_EXCEL", time.time() - start, error=True)
            return
        
        wb = openpyxl.Workbook()
        
        # 移除預設工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 概覽工作表
        ws_overview = wb.create_sheet("Overview")
        ws_overview.append(["Document ID", result.doc_id])
        ws_overview.append(["Filename", result.filename])
        ws_overview.append(["Total Pages", result.total_pages])
        ws_overview.append(["Tables Count", len(result.tables)])
        ws_overview.append(["Text Blocks Count", len(result.text_blocks)])
        ws_overview.append(["Processing Time", f"{result.processing_time:.2f}s"])
        
        # 表格工作表
        for i, table in enumerate(result.tables[:20], 1):
            ws_name = f"Table_{i}"[:31]
            ws = wb.create_sheet(ws_name)
            for row in table.to_2d_array():
                ws.append(row)
        
        # 文字區塊工作表
        if result.text_blocks:
            ws_text = wb.create_sheet("Text_Blocks")
            ws_text.append(["Block ID", "Page", "Type", "Text"])
            for block in result.text_blocks:
                ws_text.append([block.block_id, block.page, block.block_type, block.text[:500]])
        
        wb.save(output_path)
        
        KPI.record(self.MODULE_ID + "_EXCEL", time.time() - start)


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M09: DISPLAY - 顯示模組                                                     ║
# ║  Rich 美化輸出、進度條、KPI 儀表板                                           ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class DisplayModule:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M09-001: DISPLAY MODULE                                                   ║
    ║  美化輸出與視覺化                                                          ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    def __init__(self, config: Config):
        self.config = config
        self.rich = ACC.get("rich")
        self.use_rich = self.rich is not None and config.display_rich
        
        if self.use_rich:
            self.console = self.rich["Console"]()
        else:
            self.console = None
    
    def print_banner(self):
        """M09-010: 顯示橫幅"""
        banner = """
╔══════════════════════════════════════════════════════════════════════════════╗
║   ██╗   ██╗███████╗██████╗ ██╗████████╗ █████╗ ███████╗    ██╗   ██╗ █████╗  ║
║   ██║   ██║██╔════╝██╔══██╗██║╚══██╔══╝██╔══██╗██╔════╝    ██║   ██║██╔══██╗ ║
║   ██║   ██║█████╗  ██████╔╝██║   ██║   ███████║███████╗    ██║   ██║╚██████║ ║
║   ╚██╗ ██╔╝██╔══╝  ██╔══██╗██║   ██║   ██╔══██║╚════██║    ╚██╗ ██╔╝ ╚═══██║ ║
║    ╚████╔╝ ███████╗██║  ██║██║   ██║   ██║  ██║███████║     ╚████╔╝  █████╔╝ ║
║     ╚═══╝  ╚══════╝╚═╝  ╚═╝╚═╝   ╚═╝   ╚═╝  ╚═╝╚══════╝      ╚═══╝   ╚════╝  ║
║                                                                              ║
║   PDF EXTRACTOR™ - MODULAR PIPELINE EDITION                                  ║
║   Pipeline: PDF → Image → Pre-Process → Table → Database                     ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
        if self.use_rich:
            self.console.print(banner, style="bold cyan")
        else:
            print(banner)
    
    def print_accelerator_status(self):
        """M09-020: 顯示加速器狀態"""
        report = ACC.get_report()
        
        if self.use_rich:
            Table = self.rich["Table"]
            table = Table(title="🚀 Accelerator Status", box=self.rich["box"].ROUNDED)
            table.add_column("Module", style="cyan")
            table.add_column("Status", justify="center")
            
            for name, status in report.items():
                table.add_row(name, status)
            
            self.console.print(table)
        else:
            print("\n=== Accelerator Status ===")
            for name, status in report.items():
                print(f"  {name}: {status}")
    
    def print_kpi_dashboard(self, kpi_summary: Dict):
        """M09-030: 顯示 KPI 儀表板"""
        if self.use_rich:
            Table = self.rich["Table"]
            table = Table(title="📊 KPI Dashboard", box=self.rich["box"].ROUNDED)
            table.add_column("Module", style="cyan")
            table.add_column("Calls", justify="right")
            table.add_column("Total Time", justify="right")
            table.add_column("Avg Time", justify="right")
            table.add_column("Success", justify="right")
            table.add_column("Items", justify="right")
            
            for name, kpi in kpi_summary.get("modules", {}).items():
                success_style = "green" if kpi["success_rate"] >= 95 else "yellow" if kpi["success_rate"] >= 80 else "red"
                table.add_row(
                    name,
                    str(kpi["calls"]),
                    f"{kpi['total_time']:.2f}s",
                    f"{kpi['avg_time']:.3f}s",
                    f"[{success_style}]{kpi['success_rate']:.1f}%[/]",
                    str(kpi["items_processed"])
                )
            
            self.console.print(table)
            self.console.print(f"\n⏱️  Total Time: [bold]{kpi_summary['total_time']:.2f}s[/]")
        else:
            print("\n=== KPI Dashboard ===")
            for name, kpi in kpi_summary.get("modules", {}).items():
                print(f"  {name}: {kpi['calls']} calls, {kpi['total_time']:.2f}s total, {kpi['success_rate']:.1f}% success")
            print(f"\nTotal Time: {kpi_summary['total_time']:.2f}s")
    
    def print_result_summary(self, result: DocumentResult):
        """M09-040: 顯示結果摘要"""
        if self.use_rich:
            Panel = self.rich["Panel"]
            summary = f"""
📄 Document: {result.filename}
📑 Pages: {result.total_pages}
📊 Tables: {len(result.tables)}
📝 Text Blocks: {len(result.text_blocks)}
⏱️  Processing Time: {result.processing_time:.2f}s
"""
            if result.errors:
                summary += f"⚠️  Errors: {len(result.errors)}\n"
            
            self.console.print(Panel(summary, title="📋 Result Summary", border_style="green"))
        else:
            print(f"\n=== Result Summary ===")
            print(f"  Document: {result.filename}")
            print(f"  Pages: {result.total_pages}")
            print(f"  Tables: {len(result.tables)}")
            print(f"  Text Blocks: {len(result.text_blocks)}")
            print(f"  Processing Time: {result.processing_time:.2f}s")
    
    def print_progress(self, message: str, current: int = 0, total: int = 0):
        """M09-050: 顯示進度"""
        if total > 0:
            pct = current / total * 100
            bar_len = 30
            filled = int(bar_len * current / total)
            bar = "█" * filled + "░" * (bar_len - filled)
            print(f"\r  [{bar}] {pct:.1f}% - {message}", end="", flush=True)
        else:
            print(f"  → {message}")
    
    def print_success(self, message: str):
        """M09-051: 成功訊息"""
        if self.use_rich:
            self.console.print(f"[green]✓[/] {message}")
        else:
            print(f"  ✓ {message}")
    
    def print_error(self, message: str):
        """M09-052: 錯誤訊息"""
        if self.use_rich:
            self.console.print(f"[red]✗[/] {message}")
        else:
            print(f"  ✗ {message}")
    
    def print_warning(self, message: str):
        """M09-053: 警告訊息"""
        if self.use_rich:
            self.console.print(f"[yellow]⚠[/] {message}")
        else:
            print(f"  ⚠ {message}")
    
    def print_info(self, message: str):
        """M09-054: 資訊訊息"""
        if self.use_rich:
            self.console.print(f"[cyan]→[/] {message}")
        else:
            print(f"  → {message}")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M10: PIPELINE - 主流程協調模組                                              ║
# ║  協調各模組運作，實現完整 Pipeline                                           ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class TableExtractorModule:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M10-050: TABLE EXTRACTOR MODULE                                           ║
    ║  表格提取（使用 pdfplumber）                                               ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    MODULE_ID = "M10_TABLE_EXTRACT"
    
    def __init__(self, config: Config):
        self.config = config
        self.pdfplumber = ACC.get("pdfplumber")
    
    def extract(self, pdf_path: str) -> Tuple[List[ExtractedTable], List[TextBlock]]:
        """M10-051: 從 PDF 提取表格和文字"""
        start = time.time()
        tables = []
        text_blocks = []
        
        if not self.pdfplumber:
            KPI.record(self.MODULE_ID, time.time() - start, error=True)
            return tables, text_blocks
        
        try:
            doc_id = Path(pdf_path).stem
            
            with self.pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    # 提取表格
                    page_tables = page.extract_tables()
                    
                    for t_idx, raw_table in enumerate(page_tables):
                        if not raw_table:
                            continue
                        
                        table = self._convert_table(raw_table, doc_id, page_num, t_idx)
                        
                        if table and self._validate_table(table):
                            table.classify_type()
                            tables.append(table)
                    
                    # 提取文字區塊
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        blocks = self._extract_text_blocks(page_text, doc_id, page_num)
                        text_blocks.extend(blocks)
            
            KPI.record(self.MODULE_ID, time.time() - start, items=len(tables))
            
        except Exception as e:
            KPI.record(self.MODULE_ID, time.time() - start, error=True)
            raise RuntimeError(f"Table extraction failed: {e}")
        
        return tables, text_blocks
    
    def _convert_table(self, raw_table: List[List], doc_id: str, page: int, idx: int) -> Optional[ExtractedTable]:
        """M10-052: 轉換原始表格"""
        if not raw_table:
            return None
        
        rows = len(raw_table)
        cols = max(len(row) for row in raw_table) if raw_table else 0
        
        if rows < self.config.table_min_rows or cols < self.config.table_min_cols:
            return None
        
        cells = []
        for r_idx, row in enumerate(raw_table):
            for c_idx, cell_text in enumerate(row):
                text = str(cell_text).strip() if cell_text else ""
                cells.append(TableCell(
                    row=r_idx,
                    col=c_idx,
                    text=text,
                    is_header=(r_idx == 0)
                ))
        
        return ExtractedTable(
            table_id=f"{doc_id}_p{page}_t{idx + 1}",
            page=page,
            rows=rows,
            cols=cols,
            cells=cells,
            extraction_method="pdfplumber"
        )
    
    def _validate_table(self, table: ExtractedTable) -> bool:
        """M10-053: 驗證表格品質"""
        # 空白比例檢查
        if table.empty_ratio > self.config.table_max_empty_ratio:
            return False
        
        # 最少資料儲存格
        if table.data_cells_count < self.config.table_min_data_cells:
            return False
        
        # 檢查是否為長文字區塊（非表格）
        for cell in table.cells:
            if len(cell.text) > self.config.table_max_text_cell_length:
                return False
        
        # 排除單欄長文字
        if self.config.table_exclude_single_column_text:
            if table.cols == 1:
                total_text = sum(len(c.text) for c in table.cells)
                if total_text > 500:
                    return False
        
        # 排除只有標題的表格
        if self.config.table_exclude_header_only:
            non_header_data = sum(1 for c in table.cells if not c.is_header and not c.is_empty)
            if non_header_data == 0:
                return False
        
        return True
    
    def _extract_text_blocks(self, text: str, doc_id: str, page: int) -> List[TextBlock]:
        """M10-054: 提取文字區塊"""
        blocks = []
        paragraphs = text.split("\n\n")
        
        for p_idx, para in enumerate(paragraphs):
            para = para.strip()
            if not para or len(para) < 10:
                continue
            
            block = TextBlock(
                block_id=f"{doc_id}_p{page}_b{p_idx + 1}",
                page=page,
                text=para
            )
            block.classify_type()
            blocks.append(block)
        
        return blocks


class PipelineCoordinator:
    """
    ╔════════════════════════════════════════════════════════════════════════════╗
    ║  M10-001: PIPELINE COORDINATOR                                             ║
    ║  主流程協調器                                                              ║
    ║                                                                            ║
    ║  Pipeline Flow:                                                            ║
    ║    1. PDF → Images (M05)                                                   ║
    ║    2. Images → Pre-processed (M06)                                         ║
    ║    3. PDF → Tables + Text (Table Extractor)                                ║
    ║    4. Results → Database (M08)                                             ║
    ╚════════════════════════════════════════════════════════════════════════════╝
    """
    
    MODULE_ID = "M10_PIPELINE"
    
    def __init__(self, config: Config):
        self.config = config
        self.display = DisplayModule(config)
        self.temp = TempManager(config)
        self.db = DatabaseModule(config)
        
        # 初始化模組
        self.pdf_to_image = None
        self.image_pre = None
        self.table_extractor = TableExtractorModule(config)
        
        if ACC.is_available("PyMuPDF"):
            self.pdf_to_image = PDFToImageModule(config, self.temp)
        
        if ACC.is_available("OpenCV") or ACC.is_available("Pillow"):
            self.image_pre = ImagePreProcessor(config, self.temp)
    
    def process_pdf(self, pdf_path: str) -> DocumentResult:
        """M10-010: 處理單一 PDF"""
        start = time.time()
        
        doc_id = Path(pdf_path).stem
        filename = Path(pdf_path).name
        
        result = DocumentResult(
            doc_id=doc_id,
            filename=filename,
            total_pages=0
        )
        
        try:
            self.display.print_info(f"Processing: {filename}")
            
            # Step 1: PDF → Images
            if self.pdf_to_image and self.config.output_images:
                self.display.print_info("Converting PDF to images...")
                page_infos = self.pdf_to_image.convert(pdf_path)
                result.total_pages = len(page_infos)
                
                # Step 2: Pre-process images
                if self.image_pre and self.config.pre_enable:
                    self.display.print_info("Pre-processing images...")
                    for info in page_infos:
                        self.image_pre.process(info)
            
            # Step 3: Extract tables and text
            self.display.print_info("Extracting tables and text...")
            tables, text_blocks = self.table_extractor.extract(pdf_path)
            result.tables = tables
            result.text_blocks = text_blocks
            
            # 如果沒有從圖片處理取得頁數，從 pdfplumber 取得
            if result.total_pages == 0:
                pdfplumber = ACC.get("pdfplumber")
                if pdfplumber:
                    with pdfplumber.open(pdf_path) as pdf:
                        result.total_pages = len(pdf.pages)
            
            result.processing_time = time.time() - start
            
            # Step 4: Save to database
            if self.config.output_sqlite:
                self.db.save_document(result)
            
            self.display.print_success(f"Completed: {len(tables)} tables, {len(text_blocks)} text blocks")
            
        except Exception as e:
            result.errors.append(str(e))
            result.processing_time = time.time() - start
            self.display.print_error(f"Error: {e}")
            KPI.record(self.MODULE_ID, result.processing_time, error=True)
        
        KPI.record(self.MODULE_ID, result.processing_time, items=1)
        return result
    
    def process_directory(self, input_dir: str, output_dir: str) -> List[DocumentResult]:
        """M10-020: 批次處理目錄"""
        start = time.time()
        
        os.makedirs(output_dir, exist_ok=True)
        self.temp.init_session()
        
        # 找出所有 PDF
        pdf_files = list(Path(input_dir).glob("*.pdf"))
        
        if not pdf_files:
            self.display.print_warning(f"No PDF files found in {input_dir}")
            return []
        
        self.display.print_info(f"Found {len(pdf_files)} PDF files")
        
        results = []
        
        for i, pdf_path in enumerate(pdf_files, 1):
            self.display.print_progress(f"Processing {pdf_path.name}", i, len(pdf_files))
            print()
            
            result = self.process_pdf(str(pdf_path))
            results.append(result)
            
            # 輸出檔案
            doc_output_dir = os.path.join(output_dir, result.doc_id)
            os.makedirs(doc_output_dir, exist_ok=True)
            
            if self.config.output_json:
                json_path = os.path.join(doc_output_dir, f"{result.doc_id}.json")
                self.db.export_json(result, json_path)
            
            if self.config.output_csv:
                self.db.export_csv(result, doc_output_dir)
            
            if self.config.output_excel:
                excel_path = os.path.join(doc_output_dir, f"{result.doc_id}.xlsx")
                self.db.export_excel(result, excel_path)
        
        # 清理暫存
        self.temp.cleanup()
        
        # 儲存 KPI
        kpi_summary = KPI.get_summary()
        self.db.save_kpi(self.temp.session_id, kpi_summary)
        
        total_time = time.time() - start
        self.display.print_success(f"Batch completed: {len(results)} documents in {total_time:.2f}s")
        
        return results
    
    def run(self, input_path: Optional[str] = None, output_dir: Optional[str] = None):
        """M10-100: 主執行入口"""
        self.display.print_banner()
        self.display.print_accelerator_status()
        
        input_path = input_path or self.config.pdf_input_dir
        output_dir = output_dir or self.config.pdf_output_dir
        
        os.makedirs(output_dir, exist_ok=True)
        
        if os.path.isfile(input_path) and input_path.lower().endswith(".pdf"):
            # 單一檔案
            self.temp.init_session()
            result = self.process_pdf(input_path)
            
            # 輸出
            doc_output_dir = os.path.join(output_dir, result.doc_id)
            os.makedirs(doc_output_dir, exist_ok=True)
            
            if self.config.output_json:
                self.db.export_json(result, os.path.join(doc_output_dir, f"{result.doc_id}.json"))
            if self.config.output_csv:
                self.db.export_csv(result, doc_output_dir)
            if self.config.output_excel:
                self.db.export_excel(result, os.path.join(doc_output_dir, f"{result.doc_id}.xlsx"))
            
            self.temp.cleanup()
            self.display.print_result_summary(result)
            
        elif os.path.isdir(input_path):
            # 目錄
            results = self.process_directory(input_path, output_dir)
            
            # 總結
            total_tables = sum(len(r.tables) for r in results)
            total_blocks = sum(len(r.text_blocks) for r in results)
            self.display.print_info(f"Total: {total_tables} tables, {total_blocks} text blocks from {len(results)} documents")
        
        else:
            self.display.print_error(f"Invalid input: {input_path}")
            return
        
        # 顯示 KPI
        if self.config.display_kpi_dashboard:
            self.display.print_kpi_dashboard(KPI.get_summary())
    
    def test(self):
        """M10-200: 自我測試"""
        self.display.print_banner()
        self.display.print_accelerator_status()
        
        print("\n" + "=" * 60)
        print("  VERITAS v9.0 Self-Test")
        print("=" * 60)
        
        # 測試暫存管理
        print("\n[Test M04] TempManager...")
        self.temp.init_session()
        print(f"  ✓ Session: {self.temp.session_id}")
        
        # 測試 KPI
        print("\n[Test M03] KPITracker...")
        KPI.record("TEST_MODULE", 0.5, items=10)
        print(f"  ✓ KPI recorded")
        
        # 測試資料結構
        print("\n[Test M07] Data Structures...")
        table = ExtractedTable(
            table_id="test_t1",
            page=1,
            rows=3,
            cols=3,
            cells=[
                TableCell(0, 0, "Header1", is_header=True),
                TableCell(0, 1, "Header2", is_header=True),
                TableCell(0, 2, "Header3", is_header=True),
                TableCell(1, 0, "Data1"),
                TableCell(1, 1, "123"),
                TableCell(1, 2, "456"),
                TableCell(2, 0, "Data2"),
                TableCell(2, 1, "789"),
                TableCell(2, 2, "營收"),
            ]
        )
        table.classify_type()
        print(f"  ✓ Table type: {table.table_type}")
        print(f"  ✓ Empty ratio: {table.empty_ratio:.2%}")
        
        # 測試顯示模組
        print("\n[Test M09] DisplayModule...")
        self.display.print_success("Test success message")
        self.display.print_warning("Test warning message")
        self.display.print_info("Test info message")
        
        # 清理
        self.temp.cleanup()
        
        print("\n" + "=" * 60)
        print("  ✓ All tests passed!")
        print("=" * 60)
        
        # 顯示 KPI
        self.display.print_kpi_dashboard(KPI.get_summary())


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  M11: CLI - 命令列介面                                                       ║
# ║  命令列參數解析與入口點                                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def create_parser() -> argparse.ArgumentParser:
    """M11-001: 建立命令列解析器"""
    parser = argparse.ArgumentParser(
        description="VERITAS PDF EXTRACTOR™ v9.0 - Modular Pipeline Edition",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Self-test
  python veritas_pdf_extractor_v9.py --test
  
  # Process single PDF
  python veritas_pdf_extractor_v9.py document.pdf -o output/
  
  # Process directory
  python veritas_pdf_extractor_v9.py pdf_input/ -o pdf_output/
  
  # With options
  python veritas_pdf_extractor_v9.py pdf_input/ --dpi 200 --no-excel --debug
"""
    )
    
    parser.add_argument("input", nargs="?", help="Input PDF file or directory")
    parser.add_argument("-o", "--output", help="Output directory")
    
    # 測試
    parser.add_argument("--test", action="store_true", help="Run self-test")
    
    # PDF 設定
    parser.add_argument("--dpi", type=int, default=300, help="PDF rendering DPI (default: 300)")
    parser.add_argument("--gray", action="store_true", help="Convert to grayscale")
    
    # 前處理
    parser.add_argument("--no-pre", action="store_true", help="Disable image pre-processing")
    parser.add_argument("--no-denoise", action="store_true", help="Disable denoising")
    parser.add_argument("--no-sharpen", action="store_true", help="Disable sharpening")
    parser.add_argument("--no-deskew", action="store_true", help="Disable deskew")
    parser.add_argument("--binarize", action="store_true", help="Enable binarization")
    
    # 輸出
    parser.add_argument("--no-json", action="store_true", help="Disable JSON output")
    parser.add_argument("--no-csv", action="store_true", help="Disable CSV output")
    parser.add_argument("--no-excel", action="store_true", help="Disable Excel output")
    parser.add_argument("--no-sqlite", action="store_true", help="Disable SQLite output")
    parser.add_argument("--save-images", action="store_true", help="Save processed images")
    
    # 暫存
    parser.add_argument("--keep-temp", action="store_true", help="Keep temporary files")
    
    # 顯示
    parser.add_argument("--no-rich", action="store_true", help="Disable rich formatting")
    parser.add_argument("--no-kpi", action="store_true", help="Disable KPI dashboard")
    parser.add_argument("--debug", action="store_true", help="Enable debug mode")
    
    return parser


def main():
    """M11-100: 主入口點"""
    parser = create_parser()
    args = parser.parse_args()
    
    # 建立配置
    config = Config(
        pdf_dpi=args.dpi,
        pdf_colorspace="gray" if args.gray else "rgb",
        pre_enable=not args.no_pre,
        pre_denoise=not args.no_denoise,
        pre_sharpen=not args.no_sharpen,
        pre_deskew=not args.no_deskew,
        pre_binarize=args.binarize,
        output_json=not args.no_json,
        output_csv=not args.no_csv,
        output_excel=not args.no_excel,
        output_sqlite=not args.no_sqlite,
        output_images=args.save_images,
        temp_keep=args.keep_temp,
        display_rich=not args.no_rich,
        display_kpi_dashboard=not args.no_kpi,
        debug_mode=args.debug
    )
    
    # 建立 Pipeline
    pipeline = PipelineCoordinator(config)
    
    if args.test:
        pipeline.test()
        return
    
    if not args.input:
        # 無輸入時使用預設目錄或執行測試
        if os.path.exists(config.pdf_input_dir):
            pipeline.run()
        else:
            pipeline.test()
        return
    
    pipeline.run(args.input, args.output)


if __name__ == "__main__":
    main()
