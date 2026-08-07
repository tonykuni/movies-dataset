#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
VIA Project  ——  via_project.py   (\u5c08 \u00b7 \u90f5\u4ef6\u2192\u5c08\u6848\u7ba1\u7406)
================================================================================
\u64f7\u53d6\u3010\u7279\u5b9a\u671f\u9593 + \u7279\u5b9a\u4e3b\u984c\u3011\u7684\u90f5\u4ef6 \u2192 \u8f49\u6210\u4e00\u500b\u53ef\u7ba1\u7406\u7684\u5c08\u6848\u3002
\u5efa\u5728 mailforge \u60c5\u5831\u4e4b\u4e0a\uff1a
  \u2022 \u671f\u9593\u7be9\u9078\uff08\u90f5\u4ef6 Date\uff09+ \u4e3b\u984c\u7be9\u9078\uff08\u7968\u865f\u932f/\u4e3b\u65e8/\u95dc\u9375\u8a5e\uff09
  \u2022 \u6848\u4ef6\u7fa4 + \u5408\u4f75\u6642\u9593\u7dda\uff08\u90f5\u4ef6 + \u884c\u4e8b\u66c6\u4e8b\u4ef6\uff09
  \u2022 \u4efb\u52d9\u6e05\u55ae\uff08\u5f85\u56de\u8986/\u5f85\u4ea4\u4ed8/\u5230\u671f\uff09\u3001\u5229\u5bb3\u95dc\u4fc2\u4eba\u3001\u91cc\u7a0b\u7891\u3001\u72c0\u614b\u7d1a\u5225
  \u2022 \u5c08\u6848\u7de8\u865f VIA-PROJ-YYYYMMDD-######
append-only\uff1b\u7d55\u4e0d\u6367\u9020\uff1bTW \u984f\u8272\uff1b\u672c\u6a5f\u3002
================================================================================
"""
import os, sys, re, json, hashlib
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import via_forge as F
import via_mailforge as M

UTC = timezone.utc


def _parse_when(s):
    if not s:
        return None
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=UTC)
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=UTC)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=UTC)
        except Exception:
            continue
    return None


class ProjectEngine:
    def __init__(self, cfg=None):
        self.cfg = cfg or F.load_config()
        self.mail = M.MailForgeEngine(self.cfg)
        self._seq = 0

    def _proj_id(self, name):
        self._seq += 1
        day = datetime.now(UTC).strftime("%Y%m%d")
        h = hashlib.blake2s(name.encode("utf-8"), digest_size=3).hexdigest().upper()
        return "VIA-PROJ-%s-%06d-%s" % (day, self._seq, h)

    # ---------- \u671f\u9593 + \u4e3b\u984c\u7be9\u9078 ----------
    def filter_emails(self, records, topic=None, since=None, until=None):
        since_dt, until_dt = _parse_when(since), _parse_when(until)
        emails = [r for r in records if isinstance(r, dict) and r.get("status") == "OK"
                  and r.get("kind") == "EMAIL" and r.get("email_intel")]
        out = []
        for r in emails:
            hdr = (r.get("structured") or {}).get("headers", {})
            t = M.MailForgeEngine._msg_time(r)
            if since_dt and (not t or t < since_dt):
                continue
            if until_dt and (not t or t > until_dt):
                continue
            if topic:
                tk = topic.strip().lower()
                ei = r["email_intel"]
                hay = " ".join([ei.get("norm_subject", ""), ei.get("cluster_label", "").lower(),
                                ei.get("thread_key", "").lower(),
                                " ".join(w for w in (r.get("aspects") or {}).get("keywords", []))])
                # \u7968\u865f\u932f\u6216\u5b57\u4e32\u547d\u4e2d
                tkt = re.sub(r"[-\s]", "", tk)
                if tk not in hay and ("tkt:" + tkt) not in ei.get("thread_key", "").lower():
                    continue
            out.append(r)
        return out

    # ---------- \u5efa\u7acb\u5c08\u6848 ----------
    def create_project(self, name, records, topic=None, since=None, until=None, calendar=None):
        subset = self.filter_emails(records, topic=topic, since=since, until=until)
        ci = self.mail.case_intelligence(subset)
        cases = ci["cases"]

        # \u5408\u4f75\u6642\u9593\u7dda\uff1a\u90f5\u4ef6\u4e8b\u4ef6 + \u884c\u4e8b\u66c6\u4e8b\u4ef6
        timeline = []
        for c in cases:
            for ev in c.get("timeline", []):
                timeline.append({"when": ev.get("time"), "type": "EMAIL",
                                 "who": ev.get("sender"), "what": ev.get("activity"),
                                 "case": c.get("cluster_label"), "link": ev.get("link")})
        for cev in (calendar or []):
            timeline.append({"when": (_parse_when(cev.get("start")).isoformat()
                                      if _parse_when(cev.get("start")) else None),
                             "type": "CALENDAR", "who": cev.get("organizer", ""),
                             "what": cev.get("subject", ""), "case": None,
                             "location": cev.get("location", "")})
        timeline.sort(key=lambda e: (e["when"] or ""))

        # \u4efb\u52d9\u6e05\u55ae\uff1a\u5f85\u56de\u8986/\u5f85\u4ea4\u4ed8 + \u5230\u671f
        tasks = []
        for c in cases:
            fa = c["future_action"]
            if fa["needs_reply"]:
                tasks.append({"task": fa["action"], "owner": ", ".join(fa["owed_by"]) or "?",
                              "case": c["cluster_label"], "priority": fa["priority"],
                              "stale_days": fa["stale_days"], "type": "REPLY",
                              "urgency": fa["urgency_score"], "impact": fa["impact_score"]})
            if "DEADLINE" in (c.get("key_points") or {}):
                tasks.append({"task": "\u5230\u671f\u8ffd\u8e64: " + c["key_points"]["DEADLINE"][:40],
                              "owner": "\u5c08\u6848", "case": c["cluster_label"],
                              "priority": "MED", "type": "DEADLINE"})
        prio = {"CRITICAL": 0, "HIGH": 1, "MED": 2, "LOW": 3}
        tasks.sort(key=lambda t: prio.get(t.get("priority"), 9))

        # \u5229\u5bb3\u95dc\u4fc2\u4eba
        stake = self.mail.stakeholders(subset)

        # \u91cc\u7a0b\u7891\uff1acommitment + \u884c\u4e8b\u66c6\u4e8b\u4ef6
        milestones = []
        for c in cases:
            for s in c.get("solution_matrix", []):
                if s.get("activity") == "Commitment":
                    milestones.append({"milestone": "\u627f\u8afe\u4ea4\u4ed8", "case": c["cluster_label"],
                                       "who": s.get("by"), "when": s.get("time")})
        for cev in (calendar or []):
            milestones.append({"milestone": cev.get("subject", "\u884c\u4e8b\u66c6\u4e8b\u4ef6"),
                               "case": None, "who": cev.get("organizer", ""),
                               "when": cev.get("start")})

        # \u72c0\u614b\u7d1a\u5225
        crit = any(c["future_action"]["priority"] == "CRITICAL" for c in cases)
        overdue = any((t.get("stale_days") or 0) > 3 and t["type"] == "REPLY" for t in tasks)
        open_replies = sum(1 for c in cases if c["future_action"]["needs_reply"])
        status = ("BLOCKED" if crit else "AT_RISK" if overdue else
                  "ACTIVE" if open_replies else "ON_TRACK")

        return {
            "project_id": self._proj_id(name),
            "name": name, "topic": topic,
            "period": {"since": since, "until": until},
            "created_at": datetime.now(UTC).isoformat(),
            "status": status,
            "email_count": len(subset), "case_count": len(cases),
            "open_replies": open_replies,
            "cases": [{"case_id": c["case_id"], "cluster_label": c["cluster_label"],
                       "subject": c["subject"], "priority": c["future_action"]["priority"],
                       "needs_reply": c["future_action"]["needs_reply"],
                       "message_count": c["message_count"]} for c in cases],
            "timeline": timeline,
            "tasks": tasks,
            "stakeholders": stake,
            "milestones": milestones,
            "dedup_policy": "\u53bb\u91cd\u4fdd\u7559\u4e0d\u79fb\u9664\uff08\u53ea\u589e\u4e0d\u6e1b\uff09",
        }

    # ---------- \u5408\u898f\u6a4b\u63a5\uff1a\u532f\u51fa WorkMatrix xlsx\uff08\u63a5 Power Automate / Excel Online\uff09----------
    # SOP_Compliant_Automation.md \u7684\u547d\u540d\u8868\u683c schema\uff1bPower Automate \u53ea\u8a8d\u5f97\u300c\u8868\u683c\u300d\u3002
    WORKMATRIX_COLS = ["\u65e5\u671f", "\u5bc4\u4ef6\u4eba", "\u4e3b\u65e8", "\u5167\u5bb9\u6458\u8981",
                       "\u6848\u4ef6\u7de8\u865f", "\u72c0\u614b", "\u4e0b\u4e00\u6b65", "\u8cac\u4efb\u4eba"]

    def to_workmatrix_rows(self, proj):
        rows = []
        state_of = {c["case_id"]: ("\u5f85\u56de\u8986" if c["needs_reply"] else "\u5df2\u56de\u8986")
                    for c in proj["cases"]}
        subj_of = {c["cluster_label"]: c["subject"] for c in proj["cases"]}
        for e in proj["timeline"]:
            if e["type"] != "EMAIL":
                continue
            cl = e.get("case")
            rows.append([
                (e.get("when") or "")[:16].replace("T", " "),
                e.get("who", ""), subj_of.get(cl, cl or ""),
                e.get("what", ""), cl or "\u5f85\u6b78\u985e",
                "\u65b0\u6536\u5230", "", "",
            ])
        return rows

    def export_workmatrix(self, proj, path):
        try:
            import openpyxl
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.styles import Font
        except Exception:
            return {"ok": False, "reason": "openpyxl \u672a\u88dd"}
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Matrix"
        ws.append(self.WORKMATRIX_COLS)
        rows = self.to_workmatrix_rows(proj)
        if not rows:  # Power Automate \u8868\u683c\u9700\u81f3\u5c11\u4e00\u5217\u7bc4\u4f8b
            rows = [["2026-07-01 09:00", "example@corp.com", "\u7bc4\u4f8b\u4e3b\u65e8",
                     "\u7bc4\u4f8b\u6458\u8981", "PN0000", "\u65b0\u6536\u5230", "", "Chris"]]
        for r in rows:
            ws.append(r)
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True)
        last = "H%d" % (len(rows) + 1)
        tbl = Table(displayName="MatrixTable", ref="A1:%s" % last)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tbl)
        for col in "ABCDEFGH":
            ws.column_dimensions[col].width = 18
        # \u9644\u9304\uff1a\u63d0\u9192\u8868\uff08\u9644\u9304 A \u4e94\u6b04\uff09
        ws2 = wb.create_sheet("Reminders")
        ws2.append(["\u7de8\u865f", "\u4e8b\u52d9\u540d\u7a31", "\u8ffd\u8e64\u65e5\u671f", "\u72c0\u614b", "\u63d0\u9192\u5df2\u5efa"])
        for t in proj.get("tasks", []):
            if t.get("type") == "DEADLINE":
                ws2.append([t.get("case", ""), t.get("task", "")[:40], "", "\u672a\u5b8c\u6210", ""])
        for cell in ws2[1]:
            cell.font = Font(name="Arial", bold=True)
        rtbl = Table(displayName="ReminderTable", ref="A1:E%d" % max(2, ws2.max_row))
        rtbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws2.add_table(rtbl)
        wb.save(path)
        return {"ok": True, "path": path, "rows": len(rows), "table": "MatrixTable"}

    # ---------- \u5408\u898f\u63d0\u9192\uff1a\u532f\u51fa ICS\uff08\u4f60\u81ea\u5df1\u532f\u5165\uff0c\u4e0d\u5beb Outlook\uff09----------
    def export_ics(self, proj, path):
        def esc(s):
            return str(s).replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;").replace("\n", "\\n")
        lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//VIA//Project//TW", "CALSCALE:GREGORIAN"]
        n = 0
        for m in proj.get("milestones", []):
            dt = _parse_when(m.get("when"))
            if not dt:
                continue
            n += 1
            lines += [
                "BEGIN:VEVENT",
                "UID:%s-%d@via" % (proj["project_id"], n),
                "DTSTART:%s" % dt.strftime("%Y%m%dT%H%M%S"),
                "DTEND:%s" % dt.strftime("%Y%m%dT%H%M%S"),
                "SUMMARY:%s" % esc("[%s] %s" % (m.get("case") or proj["name"], m.get("milestone", ""))),
                "DESCRIPTION:%s" % esc("VIA \u5c08\u6848 %s \u00b7 %s" % (proj["name"], m.get("who", ""))),
                "BEGIN:VALARM", "TRIGGER:-PT30M", "ACTION:DISPLAY",
                "DESCRIPTION:%s" % esc(m.get("milestone", "\u63d0\u9192")), "END:VALARM",
                "END:VEVENT",
            ]
        lines.append("END:VCALENDAR")
        with open(path, "w", encoding="utf-8") as f:
            f.write("\r\n".join(lines))
        return {"ok": True, "path": path, "events": n}

    def to_html(self, proj, path=None):
        sc = {"BLOCKED": "#c96b5a", "AT_RISK": "#c4943a", "ACTIVE": "#4c78a8",
              "ON_TRACK": "#5a9e6f"}.get(proj["status"], "#6b6a66")
        pc = {"CRITICAL": "#c96b5a", "HIGH": "#c4943a", "MED": "#c9b05a", "LOW": "#5a9e6f"}
        def rows_tasks():
            r = ""
            for t in proj["tasks"]:
                c = pc.get(t.get("priority"), "#6b6a66")
                r += ('<tr><td><span class="rl" style="background:%s">%s</span></td>'
                      '<td>%s</td><td>%s</td><td class="mono" style="font-size:11px">%s</td></tr>') % (
                    c, t.get("priority", ""), t["task"], t.get("owner", ""), t.get("case", ""))
            return r or '<tr><td colspan="4" style="color:#9c9890">無待辦</td></tr>'
        def rows_tl():
            r = ""
            for e in proj["timeline"][:40]:
                w = (e["when"] or "")[:16].replace("T", " ")
                tag = "\u884c\u4e8b\u66c6" if e["type"] == "CALENDAR" else "\u90f5\u4ef6"
                r += '<tr><td class="mono" style="font-size:11px">%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                    w, tag, e.get("who", ""), e.get("what", ""))
            return r
        def rows_stk():
            return "".join('<span class="cap">%s · %d</span>' % (s["stakeholder"], s["message_count"])
                           for s in proj["stakeholders"][:15]) or "\u2014"
        html = _PROJ_HTML
        rep = {
            "__NAME__": proj["name"], "__PID__": proj["project_id"], "__SCOL__": sc,
            "__STATUS__": proj["status"], "__TOPIC__": proj.get("topic") or "(\u5168\u90e8)",
            "__SINCE__": proj["period"]["since"] or "\u2014", "__UNTIL__": proj["period"]["until"] or "\u2014",
            "__EC__": str(proj["email_count"]), "__CC__": str(proj["case_count"]),
            "__OR__": str(proj["open_replies"]),
            "__TASKS__": rows_tasks(), "__TL__": rows_tl(), "__STK__": rows_stk(),
            "__MS__": str(len(proj["milestones"])), "__TS__": proj["created_at"][:19],
        }
        for k, v in rep.items():
            html = html.replace(k, v)
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
        return html


_PROJ_HTML = """<!DOCTYPE html><html lang="zh-Hant"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>VIA \u5c08\u6848 \u00b7 __NAME__</title>
<style>@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;600&family=DM+Mono&family=Noto+Sans+TC:wght@400;500&display=swap');
:root{--bg:#f5f4f0;--sf:#fff;--bd:#dbd9d3;--ink:#1e1d1a;--sub:#6b6a66;--gr:#9c9890;--seal:#b5291a}
*{box-sizing:border-box;margin:0;padding:0}body{background:var(--bg);color:var(--ink);font-family:"DM Sans","Noto Sans TC",sans-serif;font-size:13px;line-height:1.6}
.wrap{max-width:1000px;margin:0 auto;padding:26px}.mono{font-family:"DM Mono",monospace}
h1{font-family:"Syne",sans-serif;font-weight:800;font-size:22px;display:flex;align-items:center}
h2{font-family:"Syne",sans-serif;font-weight:700;font-size:13px;margin-bottom:8px}
.eyebrow{font-family:"DM Mono",monospace;font-size:11px;letter-spacing:2px;text-transform:uppercase;color:var(--sub)}
.seal{display:inline-flex;width:30px;height:30px;align-items:center;justify-content:center;border-radius:7px;background:var(--seal);color:#fff;font-weight:900;margin-right:8px}
.strip{height:6px;border-radius:2px;margin:12px 0;background:linear-gradient(90deg,#4c78a8,#439a9a,#c9b05a,#c96b5a)}
.card{background:var(--sf);border:1px solid var(--bd);border-radius:8px;padding:14px;margin:10px 0}
table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:6px 9px;text-align:left;border-bottom:1px solid var(--bd)}
th{font-family:"DM Mono",monospace;font-size:9.5px;text-transform:uppercase;color:var(--sub)}
.rl{border-radius:9px;padding:1px 9px;font-size:9.5px;font-weight:700;color:#fff}
.cap{display:inline-block;border:1px solid var(--bd);border-radius:9px;padding:2px 9px;font-size:10.5px;margin:2px;font-family:"DM Mono",monospace}
.kpi{display:flex;gap:22px;flex-wrap:wrap;margin:6px 0}.kpi div b{font-family:"Syne",sans-serif;font-size:24px}
.foot{color:var(--gr);font-size:11px;border-top:1px solid var(--bd);padding-top:10px;margin-top:14px}</style></head>
<body><div class="wrap"><div class="eyebrow">VERITAS INTELLIGENCE SYSTEM \u00b7 PROJECT</div>
<h1><span class="seal">\u5c08</span>__NAME__</h1>
<div class="mono" style="color:var(--sub);font-size:11px">__PID__ \u00b7 \u4e3b\u984c __TOPIC__ \u00b7 \u671f\u9593 __SINCE__ ~ __UNTIL__</div>
<div class="strip"></div>
<div class="card"><div class="kpi">
  <div><b style="color:__SCOL__">__STATUS__</b><br><span class="mono" style="font-size:10px;color:var(--sub)">\u72c0\u614b</span></div>
  <div><b>__EC__</b><br><span class="mono" style="font-size:10px;color:var(--sub)">\u90f5\u4ef6</span></div>
  <div><b>__CC__</b><br><span class="mono" style="font-size:10px;color:var(--sub)">\u6848\u4ef6</span></div>
  <div><b style="color:#c96b5a">__OR__</b><br><span class="mono" style="font-size:10px;color:var(--sub)">\u5f85\u56de\u8986</span></div>
  <div><b>__MS__</b><br><span class="mono" style="font-size:10px;color:var(--sub)">\u91cc\u7a0b\u7891</span></div>
</div></div>
<div class="card"><h2>\u4efb\u52d9\u6e05\u55ae\uff08\u5148\u56de\u8986\uff09</h2>
<table><thead><tr><th>\u512a\u5148</th><th>\u4efb\u52d9</th><th>\u8ca0\u8cac/\u6b20\u56de\u8986</th><th>\u6848</th></tr></thead><tbody>__TASKS__</tbody></table></div>
<div class="card"><h2>\u5408\u4f75\u6642\u9593\u7dda\uff08\u90f5\u4ef6 + \u884c\u4e8b\u66c6\uff09</h2>
<table><thead><tr><th>\u6642\u9593</th><th>\u985e\u578b</th><th>\u4eba</th><th>\u4e8b\u4ef6</th></tr></thead><tbody>__TL__</tbody></table></div>
<div class="card"><h2>\u5229\u5bb3\u95dc\u4fc2\u4eba</h2>__STK__</div>
<div class="foot">VIA \u00b7 VeritasIntelligenceAnalytics \u00b7 \u5c08 \u00b7 \u53bb\u91cd\u4fdd\u7559\u4e0d\u79fb\u9664 \u00b7 __TS__</div>
</div></body></html>"""


# ==================================================================== SELFTEST ==
def selftest():
    import tempfile
    p, f = 0, 0
    def ck(n, c):
        nonlocal p, f
        if c: p += 1; print("  [PASS] project: %s" % n)
        else: f += 1; print("  [FAIL] project: %s" % n)

    cfg = F.load_config(); side = tempfile.mkdtemp()
    for k in ("ledger", "cache", "quarantine", "reports", "runtime", "inbox", "exports"):
        cfg["paths"][k] = os.path.join(side, k); os.makedirs(cfg["paths"][k], exist_ok=True)

    d = tempfile.mkdtemp()
    mails = [
        ("qa.li@corp.com", "PN-9942 AOI \u505c\u7dda", "<a@c>", "Mon, 06 Jul 2026 09:00:00 +0800", "\u8acb\u63d0\u4f9b\u91cd\u5de5\u65b9\u6848 \u98a8\u96aa\u9ad8"),
        ("rd.wang@corp.com", "Re: PN-9942 AOI \u505c\u7dda", "<b@c>", "Tue, 07 Jul 2026 14:00:00 +0800", "\u627f\u8afe \u9810\u8a08 07/10 \u63d0\u4f9b ECN"),
        ("vendor@acme.com", "PO-100 \u4ea4\u671f", "<c@c>", "Fri, 20 Jun 2026 09:00:00 +0800", "\u8acb\u78ba\u8a8d\u4ea4\u671f"),  # \u4e0d\u540c\u4e3b\u984c\u3001\u671f\u9593\u5916
    ]
    for i, (fr, su, mid, dt, bd) in enumerate(mails):
        open(os.path.join(d, "m%d.eml" % i), "w", encoding="utf-8").write(
            "From: %s\nTo: team@corp.com\nSubject: %s\nMessage-ID: %s\nDate: %s\n\n%s" % (fr, su, mid, dt, bd))
    recs = M.MailForgeEngine(cfg).read_and_classify(d, do_fix=True)

    pe = ProjectEngine(cfg)
    # \u4e3b\u984c\u7be9\u9078\uff1a\u53ea\u8981 PN-9942
    proj = pe.create_project("AOI \u505c\u7dda\u5c08\u6848", recs, topic="PN-9942")
    ck("project id format", proj["project_id"].startswith("VIA-PROJ-"))
    ck("topic filter keeps PN-9942 only", proj["email_count"] == 2 and proj["case_count"] == 1)
    ck("has tasks", len(proj["tasks"]) >= 1)
    ck("has stakeholders", len(proj["stakeholders"]) >= 1)
    ck("status computed", proj["status"] in ("BLOCKED", "AT_RISK", "ACTIVE", "ON_TRACK"))
    ck("milestone from commitment", any(m["milestone"] == "\u627f\u8afe\u4ea4\u4ed8" for m in proj["milestones"]))

    # \u671f\u9593\u7be9\u9078\uff1a\u53ea\u8981 7 \u6708
    proj2 = pe.create_project("\u4e03\u6708\u5c08\u6848", recs, since="2026-07-01", until="2026-07-31")
    ck("period filter excludes June", all("PO-100" not in c["subject"] for c in proj2["cases"]))

    # \u884c\u4e8b\u66c6\u5408\u4f75
    cal = [{"subject": "PN-9942 \u6aa2\u8a0e\u6703", "start": "2026-07-09 10:00:00", "organizer": "pm.chen@corp.com"}]
    proj3 = pe.create_project("\u542b\u884c\u4e8b\u66c6", recs, topic="PN-9942", calendar=cal)
    ck("calendar merged into timeline", any(e["type"] == "CALENDAR" for e in proj3["timeline"]))
    ck("calendar in milestones", any(m["when"] == "2026-07-09 10:00:00" for m in proj3["milestones"]))

    # HTML
    html = pe.to_html(proj)
    ck("project html renders", "\u5c08" in html and proj["project_id"] in html and "\u4efb\u52d9\u6e05\u55ae" in html)

    # \u5408\u898f\u6a4b\u63a5\uff1aWorkMatrix xlsx (\u547d\u540d\u8868\u683c) + ICS
    xlsx_path = os.path.join(side, "WorkMatrix.xlsx")
    wm = pe.export_workmatrix(proj3, xlsx_path)
    ck("workmatrix xlsx export", wm.get("ok") and os.path.exists(xlsx_path) and wm["table"] == "MatrixTable")
    if wm.get("ok"):
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Matrix"]
        ck("workmatrix named table present", "MatrixTable" in [t for t in ws.tables])
        ck("workmatrix schema cols", [c.value for c in ws[1]][:8] == pe.WORKMATRIX_COLS)
    ics_path = os.path.join(side, "reminders.ics")
    ic = pe.export_ics(proj3, ics_path)
    ck("ics export", ic.get("ok") and os.path.exists(ics_path))
    ics_txt = open(ics_path, encoding="utf-8").read()
    ck("ics valid VCALENDAR", "BEGIN:VCALENDAR" in ics_txt and "BEGIN:VEVENT" in ics_txt and "VALARM" in ics_txt)

    return p, f


def main():
    import argparse
    ap = argparse.ArgumentParser(description="VIA Project \u2014 \u90f5\u4ef6\u2192\u5c08\u6848")
    ap.add_argument("--scan"); ap.add_argument("--name", default="\u5c08\u6848")
    ap.add_argument("--topic"); ap.add_argument("--since"); ap.add_argument("--until")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()
    if args.selftest:
        a, b = selftest(); print("project: %d PASS / %d FAIL" % (a, b)); sys.exit(0 if b == 0 else 1)
    if args.scan:
        recs = M.MailForgeEngine().read_and_classify(args.scan, do_fix=True)
        proj = ProjectEngine().create_project(args.name, recs, topic=args.topic,
                                              since=args.since, until=args.until)
        print(json.dumps({k: v for k, v in proj.items() if k != "timeline"}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
